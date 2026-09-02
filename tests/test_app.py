"""End-to-end tests for the API and the clash-detection engine.

Run with:  python -m pytest -q
"""

from __future__ import annotations

import os
import re
import sys
from pathlib import Path

import pytest

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from timetable.config import Settings  # noqa: E402
from timetable.db import init_database  # noqa: E402
from timetable.services import Assignment, TimetableService, ValidationError, to_minutes  # noqa: E402
from timetable.web import create_app  # noqa: E402


@pytest.fixture()
def settings(tmp_path: Path) -> Settings:
    os.environ["TTG_DATA_DIR"] = str(tmp_path)
    return Settings(database_url=f"sqlite:///{(tmp_path / 'test.db').as_posix()}", log_dir=tmp_path)


@pytest.fixture()
def app(settings: Settings):
    application = create_app(settings)
    application.config["TESTING"] = True
    return application


@pytest.fixture()
def client(app):
    return app.test_client()


@pytest.fixture()
def service(settings: Settings) -> TimetableService:
    return TimetableService(init_database(settings.database_url))


# --------------------------------------------------------------------------- #
# smoke
# --------------------------------------------------------------------------- #
def test_index_renders(client):
    response = client.get("/")
    assert response.status_code == 200
    assert b"Automated Timetable Generator" in response.data


def test_health_reports_seeded_stats(client):
    payload = client.get("/api/health").get_json()
    assert payload["status"] == "ok"
    assert payload["stats"]["courses"] == 18
    assert payload["stats"]["rooms"] == 36
    assert payload["stats"]["sections"] == 45


def test_courses_include_instructor_and_headcount(client):
    courses = client.get("/api/courses").get_json()
    assert len(courses) == 65          # 45 lectures + 20 labs
    sample = courses[0]
    for key in ("id", "name", "color", "section", "instructor", "num_students"):
        assert key in sample


def test_course_details_and_404(client):
    details = client.get("/api/course-details/101/A").get_json()
    assert details["section"] == "A"
    assert isinstance(details["students"], list)
    assert client.get("/api/course-details/101/Z").status_code == 404
    assert client.get("/api/course-details/9999/A").status_code == 404


def test_rooms_are_labelled(client):
    rooms = client.get("/api/rooms").get_json()
    assert rooms[0]["label"].startswith("A-")
    assert rooms[0]["capacity"] > 0


# --------------------------------------------------------------------------- #
# validation helpers
# --------------------------------------------------------------------------- #
@pytest.mark.parametrize("value", ["8:30", "08:30", "23:59", "00:00"])
def test_valid_times(value):
    assert to_minutes(value) >= 0


@pytest.mark.parametrize("value", ["24:00", "8:70", "half past 8", "", "0830", None])
def test_invalid_times_rejected(value):
    with pytest.raises(ValidationError):
        to_minutes(value)


def test_assignment_rejects_backwards_slot():
    with pytest.raises(ValidationError):
        Assignment.from_payload(
            {"day": 1, "start_time": "10:00", "end_time": "09:00", "room_id": 1, "course_id": 101, "section": "A"}
        )


def test_assignment_rejects_bad_day():
    with pytest.raises(ValidationError):
        Assignment.from_payload(
            {"day": 9, "start_time": "09:00", "end_time": "10:00", "room_id": 1, "course_id": 101, "section": "A"}
        )


# --------------------------------------------------------------------------- #
# clash detection
# --------------------------------------------------------------------------- #
def make(day=1, start="09:00", end="10:20", room=1, course=101, section="A"):
    return Assignment(day=day, start_time=start, end_time=end, room_id=room, course_id=course, section=section)


def test_no_conflict_on_empty_grid(service):
    assert service.check_assignment(make(), others=[]) == []


def grid_from(*assignments):
    return [
        {
            "id": None,
            "day": a.day,
            "start_time": a.start_time,
            "end_time": a.end_time,
            "room_id": a.room_id,
            "course_id": a.course_id,
            "section": a.section,
            "course_name": f"Course {a.course_id}",
        }
        for a in assignments
    ]


def test_room_double_booking_detected(service):
    placed = make(course=101, section="A", room=5)
    conflicts = service.check_assignment(make(course=103, section="A", room=5), others=grid_from(placed))
    assert any(c.kind == "room" for c in conflicts)


def test_partial_overlap_detected(service):
    placed = make(course=101, section="A", room=5, start="09:00", end="10:20")
    candidate = make(course=103, section="A", room=5, start="10:00", end="11:20")
    assert any(c.kind == "room" for c in service.check_assignment(candidate, others=grid_from(placed)))


def test_back_to_back_is_allowed(service):
    placed = make(course=101, section="A", room=5, start="09:00", end="10:20")
    candidate = make(course=103, section="A", room=5, start="10:20", end="11:40")
    assert service.check_assignment(candidate, others=grid_from(placed)) == []


def test_different_day_is_allowed(service):
    placed = make(day=1, room=5)
    candidate = make(day=2, room=5, course=103)
    assert service.check_assignment(candidate, others=grid_from(placed)) == []


def test_instructor_double_booking_detected(service):
    # Dr. Hammad Majeed (111) teaches MLOps A and MLOps B.
    placed = make(course=101, section="A", room=1)
    candidate = make(course=101, section="B", room=2)
    conflicts = service.check_assignment(candidate, others=grid_from(placed))
    assert any(c.kind == "instructor" for c in conflicts)


def test_student_clash_lists_roll_numbers(service):
    # 20I-0546 is enrolled in MLOps A (101) and Network Security A (102).
    placed = make(course=101, section="A", room=1)
    candidate = make(course=102, section="A", room=2)
    conflicts = service.check_assignment(candidate, others=grid_from(placed))
    student = [c for c in conflicts if c.kind == "student"]
    assert student, "expected a student clash"
    assert "20I-0546" in student[0].details["roll_numbers"]


def test_duplicate_section_detected(service):
    placed = make(course=101, section="A", room=1)
    candidate = make(course=101, section="A", room=2)
    assert any(c.kind == "duplicate" for c in service.check_assignment(candidate, others=grid_from(placed)))


def test_unknown_course_or_room_rejected(service):
    assert service.check_assignment(make(course=999), others=[])[0].kind == "unknown"
    assert service.check_assignment(make(room=9999), others=[])[0].kind == "unknown"


# --------------------------------------------------------------------------- #
# persistence
# --------------------------------------------------------------------------- #
def test_save_load_and_reset_roundtrip(client):
    payload = {
        "assignments": [
            {"day": 1, "start_time": "08:30", "end_time": "09:50", "room_id": 1, "course_id": 101, "section": "A"},
            {"day": 1, "start_time": "10:00", "end_time": "11:20", "room_id": 2, "course_id": 103, "section": "B"},
        ]
    }
    saved = client.post("/api/timetable", json=payload)
    assert saved.status_code == 200, saved.get_json()
    assert saved.get_json()["saved"] == 2

    entries = client.get("/api/timetable").get_json()["entries"]
    assert len(entries) == 2
    assert entries[0]["room_label"].startswith("A-")

    # saving again must replace, not duplicate
    client.post("/api/timetable", json=payload)
    assert len(client.get("/api/timetable").get_json()["entries"]) == 2

    reset = client.post("/api/timetable/reset").get_json()
    assert reset["removed"] == 2
    assert client.get("/api/timetable").get_json()["entries"] == []


def test_save_is_rejected_when_the_grid_clashes(client):
    payload = {
        "assignments": [
            {"day": 1, "start_time": "08:30", "end_time": "09:50", "room_id": 1, "course_id": 101, "section": "A"},
            {"day": 1, "start_time": "08:30", "end_time": "09:50", "room_id": 1, "course_id": 103, "section": "A"},
        ]
    }
    response = client.post("/api/timetable", json=payload)
    assert response.status_code == 409
    assert response.get_json()["saved"] == 0
    assert client.get("/api/timetable").get_json()["entries"] == []   # nothing partially written


def test_validate_endpoint_reports_candidate_conflicts(client):
    body = {
        "candidate": {"day": 1, "start_time": "08:30", "end_time": "09:50", "room_id": 1, "course_id": 101, "section": "B"},
        "grid": [
            {"day": 1, "start_time": "08:30", "end_time": "09:50", "room_id": 2, "course_id": 101, "section": "A"}
        ],
    }
    result = client.post("/api/timetable/validate", json=body).get_json()
    assert result["ok"] is False
    assert any(c["kind"] == "instructor" for c in result["conflicts"])


def test_malformed_payloads_return_400_not_500(client):
    assert client.post("/api/timetable", data="not json", content_type="application/json").status_code == 400
    assert client.post("/api/timetable", json={"assignments": [{"day": 1}]}).status_code == 400
    assert client.post("/api/timetable", json={"assignments": "nope"}).status_code == 400


def test_sql_injection_attempt_is_harmless(client):
    payload = {
        "assignments": [
            {
                "day": 1, "start_time": "08:30", "end_time": "09:50", "room_id": 1,
                "course_id": 101, "section": "A'); DROP TABLE courses;--",
            }
        ]
    }
    response = client.post("/api/timetable", json=payload)
    assert response.status_code == 409          # rejected as an unknown section
    assert client.get("/api/health").get_json()["stats"]["courses"] == 18   # table intact


def test_settings_roundtrip(client):
    client.post("/api/settings", json={"days": 6, "start": "09:00"})
    assert client.get("/api/settings").get_json()["days"] == 6


def test_legacy_routes_still_answer(client):
    assert client.post("/save-timetable", json={"assigned_courses": []}).status_code == 200
    assert client.post("/reset-timetable", json={}).status_code == 200


def test_unknown_api_route_returns_json_404(client):
    response = client.get("/api/does-not-exist")
    assert response.status_code == 404
    assert response.is_json


def test_database_failure_is_reported_gracefully(tmp_path):
    broken = Settings(database_url="postgresql+psycopg2://nobody@127.0.0.1:1/none", log_dir=tmp_path)
    application = create_app(broken)
    test_client = application.test_client()
    assert test_client.get("/").status_code == 200          # UI still loads
    assert test_client.get("/api/health").status_code == 503
    assert test_client.get("/api/courses").status_code == 503


# =========================================================================== #
# v2.1 - catalogue management, shifts, auto-fill, Excel export
# =========================================================================== #
from timetable.catalog import CatalogService  # noqa: E402
from timetable.exporters import build_workbook, format_time_range, openpyxl_available  # noqa: E402


@pytest.fixture()
def catalog(settings: Settings) -> CatalogService:
    return CatalogService(init_database(settings.database_url))


# ------------------------------- teachers ---------------------------------- #
def test_add_edit_delete_teacher(client):
    created = client.post("/api/instructors", json={
        "name": "Dr. New Person", "email": "new@uni.edu", "department": "CS", "shift": "evening"
    })
    assert created.status_code == 201
    teacher_id = created.get_json()["id"]

    listing = client.get("/api/instructors").get_json()
    assert any(t["id"] == teacher_id and t["shift"] == "evening" for t in listing)

    updated = client.put(f"/api/instructors/{teacher_id}", json={"name": "Dr. Renamed", "department": "SE"})
    assert updated.get_json()["name"] == "Dr. Renamed"

    assert client.delete(f"/api/instructors/{teacher_id}").status_code == 200
    assert not any(t["id"] == teacher_id for t in client.get("/api/instructors").get_json())


def test_teacher_validation(client):
    assert client.post("/api/instructors", json={"name": "  "}).status_code == 400
    assert client.post("/api/instructors", json={"name": "X", "email": "not-an-email"}).status_code == 400
    client.post("/api/instructors", json={"name": "Unique Person"})
    assert client.post("/api/instructors", json={"name": "unique person"}).status_code == 400  # case-insensitive


def test_busy_teacher_cannot_be_deleted(client):
    teachers = client.get("/api/instructors").get_json()
    busy = next(t for t in teachers if t["sections"] > 0)
    response = client.delete(f"/api/instructors/{busy['id']}")
    assert response.status_code == 400
    assert "section" in response.get_json()["message"].lower()


# --------------------------------- rooms ----------------------------------- #
def test_add_room_creating_its_building_on_the_fly(client):
    response = client.post("/api/rooms", json={
        "room_number": "Z-9", "building_name": "New Block", "capacity": 35, "room_type": "Lab"
    })
    assert response.status_code == 201
    room_id = response.get_json()["id"]

    rooms = client.get("/api/rooms").get_json()
    room = next(r for r in rooms if r["id"] == room_id)
    assert room["room_type"] == "Lab" and room["capacity"] == 35
    assert any(b["name"] == "New Block" for b in client.get("/api/buildings").get_json())

    assert client.delete(f"/api/rooms/{room_id}").status_code == 200


def test_duplicate_room_in_same_building_rejected(client):
    first = client.post("/api/rooms", json={"room_number": "777", "building_name": "A"})
    assert first.status_code == 201
    assert client.post("/api/rooms", json={"room_number": "777", "building_name": "A"}).status_code == 400
    # ... but the same number in another building is fine
    assert client.post("/api/rooms", json={"room_number": "777", "building_name": "B"}).status_code == 201


def test_room_used_by_saved_timetable_cannot_be_deleted(client):
    client.post("/api/timetable", json={"assignments": [
        {"day": 1, "start_time": "08:30", "end_time": "09:50", "room_id": 1, "course_id": 101, "section": "A"}
    ]})
    assert client.delete("/api/rooms/1").status_code == 400


def test_invalid_room_payloads(client):
    assert client.post("/api/rooms", json={"room_number": "", "building_name": "A"}).status_code == 400
    assert client.post("/api/rooms", json={"room_number": "1", "building_name": "A", "capacity": 0}).status_code == 400
    assert client.post("/api/rooms", json={"room_number": "1", "building_name": "A", "room_type": "Cave"}).status_code == 400


# ----------------------------- courses & codes ------------------------------ #
def test_add_course_with_code_and_sections(client):
    teacher_id = client.get("/api/instructors").get_json()[0]["id"]
    response = client.post("/api/courses", json={
        "code": "cs5099", "name": "Advanced Topics", "department": "Computer Science",
        "credit_hours": 4, "color": "#123456",
        "sections": [{"section": "a", "instructor_id": teacher_id}, {"section": "b"}],
    })
    assert response.status_code == 201
    body = response.get_json()
    assert body["code"] == "CS5099"          # normalised to upper case

    admin = next(c for c in client.get("/api/admin/courses").get_json() if c["id"] == body["id"])
    assert {s["section"] for s in admin["sections"]} == {"A", "B"}
    assert admin["sections"][0]["instructor_id"] == teacher_id

    catalogue = client.get("/api/courses").get_json()
    assert any(c["code"] == "CS5099" and c["credit_hours"] == 4 for c in catalogue)


def test_course_code_must_be_unique_and_valid(client):
    client.post("/api/courses", json={"code": "CS7000", "name": "One"})
    assert client.post("/api/courses", json={"code": "cs7000", "name": "Two"}).status_code == 400
    assert client.post("/api/courses", json={"code": "", "name": "Three"}).status_code == 400
    assert client.post("/api/courses", json={"code": "BAD/CODE", "name": "Four"}).status_code == 400
    assert client.post("/api/courses", json={"code": "OK1", "name": "", }).status_code == 400
    assert client.post("/api/courses", json={"code": "OK2", "name": "Five", "color": "red"}).status_code == 400


def test_course_gets_an_automatic_colour(client):
    body = client.post("/api/courses", json={"code": "CS7100", "name": "Auto Colour"}).get_json()
    assert re.match(r"^#[0-9A-Fa-f]{6}$", body["color"])


def test_sections_can_be_added_and_removed(client):
    course_id = client.post("/api/courses", json={"code": "CS7200", "name": "Sectioned"}).get_json()["id"]
    assert client.post(f"/api/courses/{course_id}/sections", json={"section": "z"}).status_code == 201
    admin = next(c for c in client.get("/api/admin/courses").get_json() if c["id"] == course_id)
    assert [s["section"] for s in admin["sections"]] == ["Z"]
    assert client.delete(f"/api/courses/{course_id}/sections/Z").status_code == 200
    assert client.delete(f"/api/courses/{course_id}/sections/Z").status_code == 400   # already gone


def test_course_in_saved_timetable_cannot_be_deleted(client):
    client.post("/api/timetable", json={"assignments": [
        {"day": 1, "start_time": "08:30", "end_time": "09:50", "room_id": 1, "course_id": 101, "section": "A"}
    ]})
    assert client.delete("/api/courses/101").status_code == 400
    client.post("/api/timetable/reset")
    assert client.delete("/api/courses/101").status_code == 200


# --------------------------------- shifts ----------------------------------- #
def test_shift_is_persisted_with_each_class(client):
    payload = {"assignments": [
        {"day": 1, "start_time": "08:30", "end_time": "09:50", "room_id": 1,
         "course_id": 101, "section": "A", "shift": "morning"},
        {"day": 7, "start_time": "18:00", "end_time": "19:20", "room_id": 2,
         "course_id": 103, "section": "A", "shift": "evening"},
    ]}
    assert client.post("/api/timetable", json=payload).status_code == 200
    entries = client.get("/api/timetable").get_json()["entries"]
    assert {e["shift"] for e in entries} == {"morning", "evening"}
    assert max(e["day"] for e in entries) == 7          # full 7-day week supported


def test_unknown_shift_falls_back_to_morning(client):
    client.post("/api/timetable", json={"assignments": [
        {"day": 1, "start_time": "08:30", "end_time": "09:50", "room_id": 1,
         "course_id": 101, "section": "A", "shift": "nonsense"}
    ]})
    assert client.get("/api/timetable").get_json()["entries"][0]["shift"] == "morning"


# -------------------------------- auto-fill ---------------------------------- #
def test_autofill_places_sections_without_any_clash(client):
    slots = [{"start": "08:30", "end": "09:50"}, {"start": "10:00", "end": "11:20"},
             {"start": "11:30", "end": "12:50"}]
    response = client.post("/api/timetable/autofill", json={
        "assignments": [], "days": 5, "slots": slots,
        "room_ids": [1, 2, 3, 4, 5, 6], "shift": "morning",
    })
    created = response.get_json()["created"]
    assert len(created) == 65                          # every lecture *and* every lab
    assert all(c["shift"] == "morning" for c in created)

    verdict = client.post("/api/timetable/validate", json={"assignments": created}).get_json()
    assert verdict["ok"] is True, verdict.get("reports")


def test_autofill_respects_what_is_already_placed(client):
    existing = [{"day": 1, "start_time": "08:30", "end_time": "09:50", "room_id": 1,
                 "course_id": 101, "section": "A", "shift": "morning"}]
    created = client.post("/api/timetable/autofill", json={
        "assignments": existing, "days": 5,
        "slots": [{"start": "08:30", "end": "09:50"}, {"start": "10:00", "end": "11:20"}],
        "room_ids": [1, 2, 3, 4],
    }).get_json()["created"]
    assert not any(c["course_id"] == 101 and c["section"] == "A" for c in created)
    combined = client.post("/api/timetable/validate", json={"assignments": existing + created}).get_json()
    assert combined["ok"] is True


def test_autofill_needs_a_grid(client):
    assert client.post("/api/timetable/autofill", json={"days": 5, "slots": [], "room_ids": []}).status_code == 400


# ------------------------------ Excel export --------------------------------- #
def _header_row(sheet, first_heading="Days"):
    """Find the header row under the metadata title block (it moves with the
    document identity fields, so tests must not hard-code a row number)."""
    for row in range(1, min(sheet.max_row, 20) + 1):
        if sheet.cell(row=row, column=1).value == first_heading:
            return row
    raise AssertionError(f"header row ({first_heading!r}) not found in {sheet.title}")


@pytest.mark.skipif(not openpyxl_available(), reason="openpyxl not installed")
def test_excel_export_puts_every_semester_on_its_own_sheet(client):
    client.post("/api/timetable", json={"assignments": [
        {"day": 1, "start_time": "08:30", "end_time": "09:50", "room_id": 1,
         "course_id": 101, "section": "A", "shift": "morning"},
        {"day": 6, "start_time": "18:00", "end_time": "19:20", "room_id": 2,
         "course_id": 103, "section": "A", "shift": "evening"},
    ]})
    response = client.post("/api/export/xlsx", json={"days": 7, "shift": "all"})
    assert response.status_code == 200
    assert response.headers["Content-Type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml"
    )

    import io as _io
    import openpyxl

    workbook = openpyxl.load_workbook(_io.BytesIO(response.data))
    names = workbook.sheetnames

    # a Contents page first, then Summary, then one sheet per semester ...
    assert names[0] == "Contents"
    assert names[1] == "Summary"
    assert [n for n in names if n.startswith("Semester ")] == ["Semester 1", "Semester 3"]
    # ... one sheet per weekday ...
    assert names[names.index("Semester 3") + 1:names.index("By Teacher")] == [
        "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday",
    ]
    # ... and the roll-ups the previous releases had.
    for expected in ("By Teacher", "Credit Hour Audit", "Dashboard", "Room Utilisation",
                     "Teacher Workload", "Conflict Report", "Master Data", "Unscheduled"):
        assert expected in names, expected

    summary = workbook["Summary"]
    header = _header_row(summary, "Day")
    assert [summary.cell(row=header, column=c).value for c in range(1, 6)] == [
        "Day", "Shift", "Semester", "Type", "C.Hrs",
    ]
    assert summary.cell(row=header + 1, column=1).value == "Monday"
    assert summary.auto_filter.ref is not None        # filterable

    # every semester sheet is drawn in the printed Class Schedule arrangement
    for name, expected_day in (("Semester 1", "Monday"), ("Semester 3", "Saturday")):
        semester = workbook[name]
        row = _header_row(semester)
        assert [semester.cell(row=row, column=c).value for c in range(1, 9)] == [
            "Days", "Course Code", "Course Title", "C.Hrs", "Total No.of Students",
            "Teacher's Name", "Time", "Room No",
        ]
        assert semester.cell(row=row + 1, column=1).value == expected_day
        assert "AM" in str(semester.cell(row=row + 1, column=7).value) or \
            "PM" in str(semester.cell(row=row + 1, column=7).value)

    gaps = workbook["Unscheduled"]
    gaps_header = _header_row(gaps, "Semester")
    assert gaps.cell(row=gaps_header, column=1).value == "Semester"
    assert gaps.cell(row=gaps_header + 1, column=5).value in ("Theory", "Lab")


@pytest.mark.skipif(not openpyxl_available(), reason="openpyxl not installed")
def test_excel_export_works_on_an_unsaved_grid(client):
    response = client.post("/api/export/xlsx", json={
        "days": 2,
        "assignments": [{"day": 2, "start_time": "09:00", "end_time": "10:00", "room_id": 5,
                         "course_id": 104, "section": "A", "shift": "morning"}],
    })
    assert response.status_code == 200
    assert client.get("/api/timetable").get_json()["entries"] == []      # nothing was saved


def test_excel_export_rejects_an_empty_timetable(client):
    assert client.post("/api/export/xlsx", json={}).status_code == 400


# --------------------- exports written to a chosen folder --------------------- #
def _one_class(client):
    client.post("/api/timetable", json={"assignments": [
        {"day": 1, "start_time": "08:30", "end_time": "09:50", "room_id": 1,
         "course_id": 101, "section": "A", "shift": "morning"},
    ]})


@pytest.mark.skipif(not openpyxl_available(), reason="openpyxl not installed")
def test_export_is_written_into_the_requested_folder(client, tmp_path):
    """Exports land beside the project instead of in the Downloads folder."""
    _one_class(client)
    folder = tmp_path / "Spring 2026"

    response = client.post("/api/export/xlsx", json={"folder": str(folder)})
    assert response.status_code == 200
    payload = response.get_json()
    assert payload["saved"] is True
    written = Path(payload["path"])
    assert written.parent == folder.resolve()      # folder was created for us
    assert written.is_file() and written.stat().st_size > 0
    assert not list(folder.glob("*.part"))         # atomic write left no debris


@pytest.mark.skipif(not openpyxl_available(), reason="openpyxl not installed")
def test_repeated_exports_are_versioned_instead_of_numbered(client, tmp_path):
    """Three exports become rev1/rev2/rev3 - a name that says which is current,
    rather than "timetable (2).xlsx" which says nothing at all."""
    _one_class(client)
    names = [
        Path(client.post("/api/export/xlsx", json={"folder": str(tmp_path)}).get_json()["path"]).name
        for _ in range(3)
    ]
    assert names == ["timetable-rev1.xlsx", "timetable-rev2.xlsx", "timetable-rev3.xlsx"]
    assert len(list(tmp_path.glob("*.xlsx"))) == 3          # nothing was overwritten

    # The document name becomes the stem, so one folder can hold several terms.
    named = client.post("/api/export/xlsx", json={
        "folder": str(tmp_path), "document_name": "Spring 2026"}).get_json()
    assert Path(named["path"]).name == "Spring 2026-rev1.xlsx"

    # Turning versioning off brings back the old Explorer-style numbering.
    plain = client.post("/api/export/xlsx", json={
        "folder": str(tmp_path), "versioned": False}).get_json()
    assert Path(plain["path"]).name == "timetable.xlsx"
    plain_again = client.post("/api/export/xlsx", json={
        "folder": str(tmp_path), "versioned": False}).get_json()
    assert Path(plain_again["path"]).name == "timetable (2).xlsx"


@pytest.mark.skipif(not openpyxl_available(), reason="openpyxl not installed")
def test_a_new_revision_says_what_changed_since_the_last_one(client, tmp_path):
    _one_class(client)
    client.post("/api/export/xlsx", json={"folder": str(tmp_path), "document_name": "Spring 2026"})

    # move the class to another day, then export again
    client.post("/api/timetable", json={"assignments": [
        {"day": 3, "start_time": "08:30", "end_time": "09:50", "room_id": 2,
         "course_id": 101, "section": "A", "shift": "morning"},
    ]})
    second = client.post("/api/export/xlsx", json={
        "folder": str(tmp_path), "document_name": "Spring 2026"}).get_json()
    assert Path(second["path"]).name == "Spring 2026-rev2.xlsx"

    from openpyxl import load_workbook

    book = load_workbook(Path(second["path"]))
    assert book.sheetnames[0] == "Contents"
    assert book.sheetnames[1] == "Revisions"
    revisions = book["Revisions"]
    assert revisions["A1"].value.endswith("Revision 2")

    text = "\n".join(
        str(cell.value)
        for row in revisions.iter_rows()
        for cell in row
        if cell.value is not None
    )
    assert "What changed since Spring 2026-rev1.xlsx" in text
    assert "Moved" in text                       # the class changed day
    assert "rev 1" in text and "rev 2" in text   # both revisions are listed


def test_csv_export_matches_the_workbook_columns(client, tmp_path):
    _one_class(client)
    payload = client.post("/api/export/csv", json={"folder": str(tmp_path)}).get_json()
    text = Path(payload["path"]).read_text(encoding="utf-8")
    assert text.startswith("\ufeff")             # Excel-friendly BOM
    header = text.splitlines()[0]
    for column in ("Day", "Room", "Building", "Room type", "Capacity",
                   "Code", "Course", "Section", "Kind", "Teacher", "Semester", "Students"):
        assert f'"{column}"' in header
    row = text.splitlines()[1]
    assert '"Monday"' in row and '"Theory"' in row
    assert '"A-108"' in row or '"A-101"' in row   # a real room label, not an id


def test_csv_export_streams_when_no_folder_is_given(client):
    _one_class(client)
    response = client.post("/api/export/csv", json={})
    assert response.status_code == 200
    assert response.headers["Content-Type"].startswith("text/csv")
    assert response.data.startswith("\ufeff".encode("utf-8"))


def test_pdf_and_ics_also_honour_the_folder(client, tmp_path):
    _one_class(client)
    pdf = client.post("/api/publish/pdf", json={"folder": str(tmp_path), "scope": "all"}).get_json()
    assert Path(pdf["path"]).suffix == ".pdf"
    assert Path(pdf["path"]).read_bytes().startswith(b"%PDF")

    ics = client.post("/api/publish/ics", json={"folder": str(tmp_path)}).get_json()
    assert Path(ics["path"]).suffix == ".ics"
    assert "BEGIN:VCALENDAR" in Path(ics["path"]).read_text(encoding="utf-8")


@pytest.mark.skipif(
    sys.platform.startswith("win") or (hasattr(os, "geteuid") and os.geteuid() == 0),
    reason="POSIX permission bits: chmod does not restrict directories on Windows, and root ignores them",
)
def test_export_to_a_read_only_folder_explains_itself(client, tmp_path):
    _one_class(client)
    locked = tmp_path / "locked"
    locked.mkdir()
    locked.chmod(0o500)
    try:
        response = client.post("/api/export/csv", json={"folder": str(locked)})
        assert response.status_code == 400
        assert "read-only" in response.get_json()["message"].lower()
    finally:
        locked.chmod(0o700)


def test_export_rejects_a_dangerous_filename(client, tmp_path):
    _one_class(client)
    response = client.post("/api/export/csv", json={"folder": str(tmp_path), "filename": "../escape.csv"})
    assert response.status_code == 400
    assert not (tmp_path.parent / "escape.csv").exists()


# ------------------------------- migrations ---------------------------------- #
def test_migration_adds_columns_to_an_old_database(tmp_path):
    """Simulate a v2.0 database and prove it upgrades in place."""
    from sqlalchemy import create_engine, inspect, text

    db_file = tmp_path / "old.db"
    engine = create_engine(f"sqlite:///{db_file.as_posix()}")
    with engine.begin() as conn:
        conn.execute(text(
            "CREATE TABLE courses (id INTEGER PRIMARY KEY, name VARCHAR(255) NOT NULL, "
            "color VARCHAR(20) NOT NULL DEFAULT '#4c5caf', department VARCHAR(100) NOT NULL DEFAULT 'General')"
        ))
        conn.execute(text("INSERT INTO courses (id, name, department) VALUES (501, 'Legacy Course', 'Computer Science')"))
    engine.dispose()

    upgraded = init_database(f"sqlite:///{db_file.as_posix()}", seed=False)
    columns = {c["name"] for c in inspect(upgraded).get_columns("courses")}
    assert {"code", "credit_hours"} <= columns

    with upgraded.connect() as conn:
        code = conn.execute(text("SELECT code FROM courses WHERE id = 501")).scalar()
    assert code                     # a code was back-filled, not left blank


# ============================================================================ #
# v2.2 - capacity warnings, Excel import, PDF / iCalendar publishing
# ============================================================================ #
from datetime import date  # noqa: E402

from timetable.importers import build_template, import_workbook  # noqa: E402
from timetable.importers import openpyxl_available as import_openpyxl  # noqa: E402
from timetable.publishing import (  # noqa: E402
    PdfCanvas,
    build_ics,
    build_pdf,
    filter_entries,
    text_width,
)


def _sample_entries():
    return [
        {
            "id": 1, "day": 1, "start_time": "08:30", "end_time": "09:50", "shift": "morning",
            "room_id": 1, "room_label": "A-101", "course_id": 101, "code": "CS3009",
            "course_name": "Artificial Intelligence", "section": "A", "color": "#a9d2e1",
            "instructor": "Mr. Saad Salman", "num_students": 42,
        },
        {
            "id": 2, "day": 3, "start_time": "13:30", "end_time": "14:50", "shift": "evening",
            "room_id": 2, "room_label": "B-201", "course_id": 102, "code": "CS4001",
            "course_name": "Machine Learning", "section": "B", "color": "#2b3465",
            "instructor": "Dr. Aftab Maroof", "num_students": 30,
        },
    ]


_SAMPLE_ROOMS = [
    {"id": 1, "room_number": "101", "label": "A-101", "capacity": 60, "room_type": "Classroom"},
    {"id": 2, "room_number": "201", "label": "B-201", "capacity": 30, "room_type": "Lab"},
]


# ------------------------------ capacity ------------------------------------ #
def test_capacity_warning_is_raised_but_not_blocking(service):
    """A too-small room must warn, never stop the user from saving."""
    from sqlalchemy import insert, select, text as sql_text
    from timetable.db import enrollments, rooms, students

    with service.engine.begin() as conn:
        room_id = conn.execute(select(rooms.c.id)).scalar()
        conn.execute(rooms.update().where(rooms.c.id == room_id).values(capacity=1))
        roll = conn.execute(select(students.c.roll_number)).scalar()
        already = conn.execute(
            select(enrollments.c.roll_number).where(
                enrollments.c.course_id == 101, enrollments.c.section == "A",
                enrollments.c.roll_number == roll,
            )
        ).first()
        if not already:
            conn.execute(insert(enrollments).values(roll_number=roll, course_id=101, section="A"))
        conn.execute(sql_text("SELECT 1"))

    candidate = Assignment(day=1, start_time="09:00", end_time="10:20", room_id=room_id,
                           course_id=101, section="A")
    conflicts = service.check_assignment(candidate)
    capacity = [c for c in conflicts if c.kind == "capacity"]
    assert capacity, "an over-full room must produce a capacity conflict"
    assert capacity[0].severity == "warning"
    assert not any(c.severity == "error" for c in conflicts)

    result = service.save_timetable([candidate])
    assert result["ok"] and result["saved"] == 1        # warnings never block a save


def test_capacity_warning_survives_the_http_layer(client):
    from sqlalchemy import select
    from timetable.db import rooms

    engine = client.application.extensions["engine"]
    with engine.begin() as conn:
        room_id = conn.execute(select(rooms.c.id)).scalar()
        conn.execute(rooms.update().where(rooms.c.id == room_id).values(capacity=1))

    response = client.post("/api/timetable/validate", json={
        "candidate": {"day": 1, "start_time": "09:00", "end_time": "10:20", "room_id": room_id,
                      "course_id": 101, "section": "A"},
        "grid": [],
    })
    payload = response.get_json()
    assert payload["ok"] is True                        # still placeable
    assert any(c["kind"] == "capacity" for c in payload["conflicts"])


# -------------------------------- import ------------------------------------ #
@pytest.mark.skipif(not import_openpyxl(), reason="openpyxl not installed")
def test_import_template_has_every_sheet(client):
    import io as _io
    from openpyxl import load_workbook

    response = client.get("/api/import/template")
    assert response.status_code == 200
    book = load_workbook(_io.BytesIO(response.data))
    assert {"Teachers", "Buildings", "Rooms", "Courses", "Sections"} <= set(book.sheetnames)


@pytest.mark.skipif(not import_openpyxl(), reason="openpyxl not installed")
def test_import_creates_updates_and_reports_bad_rows(app):
    import io as _io
    from openpyxl import load_workbook

    catalog = app.extensions["catalog"]
    book = load_workbook(_io.BytesIO(build_template()))
    for sheet in ("Teachers", "Buildings", "Rooms", "Courses", "Sections"):
        book[sheet].delete_rows(2, book[sheet].max_row)      # drop the examples
    book["Teachers"].append(["Prof. Test Import", "t.import@uni.edu", "Physics", "evening"])
    book["Buildings"].append(["Zeta"])
    book["Rooms"].append(["Z-1", "Zeta", 25, "Lab"])
    book["Rooms"].append(["", "Zeta", 25, "Lab"])            # invalid: no room number
    book["Courses"].append(["PH1001", "Quantum Mechanics", "Physics", 4, "#C7E1C0"])
    book["Sections"].append(["PH1001", "A", "t.import@uni.edu"])
    book["Sections"].append(["PH1001", "B", "Ghost Teacher"])  # invalid: unknown teacher
    buffer = _io.BytesIO()
    book.save(buffer)

    report = import_workbook(catalog, buffer.getvalue())
    assert report["created"] == {"Teachers": 1, "Buildings": 1, "Rooms": 1, "Courses": 1, "Sections": 1}
    assert report["skipped"] == 2
    assert {e["sheet"] for e in report["errors"]} == {"Rooms", "Sections"}
    assert report["ok"] is False

    # the good rows really landed
    names = {t["name"] for t in catalog.list_instructors()}
    assert "Prof. Test Import" in names
    codes = {c["code"] for c in catalog.list_courses_admin()}
    assert "PH1001" in codes

    # re-importing the same file updates instead of duplicating
    again = import_workbook(catalog, buffer.getvalue())
    assert again["total_created"] == 0
    assert again["total_updated"] >= 4
    assert len([t for t in catalog.list_instructors() if t["name"] == "Prof. Test Import"]) == 1


@pytest.mark.skipif(not import_openpyxl(), reason="openpyxl not installed")
def test_import_rejects_a_workbook_without_known_sheets(app):
    import io as _io
    from openpyxl import Workbook

    book = Workbook()
    book.active.title = "Random"
    buffer = _io.BytesIO()
    book.save(buffer)
    with pytest.raises(ValidationError):
        import_workbook(app.extensions["catalog"], buffer.getvalue())


def test_import_rejects_a_non_excel_file(client):
    response = client.post("/api/import/xlsx", data=b"this is not a spreadsheet",
                           content_type="application/octet-stream")
    assert response.status_code == 400
    assert "Excel" in response.get_json()["message"]


def test_import_without_a_file_is_a_clean_400(client):
    assert client.post("/api/import/xlsx", data=b"").status_code == 400


# ------------------------------- publishing ---------------------------------- #
def test_pdf_writer_produces_a_valid_document():
    pdf = PdfCanvas()
    pdf.text(20, 20, "Hello")
    pdf.new_page()
    pdf.rect(10, 10, 50, 20, fill=(1, 0, 0))
    data = pdf.build("Test")
    assert data.startswith(b"%PDF-1.4")
    assert data.rstrip().endswith(b"%%EOF")
    assert data.count(b"/Type /Page ") == 2
    assert b"startxref" in data


def test_text_width_is_font_aware():
    assert text_width("iii", 10) < text_width("WWW", 10)
    assert text_width("Hello", 20) == pytest.approx(2 * text_width("Hello", 10))


@pytest.mark.parametrize("scope,pages", [("all", 5), ("teacher", 2), ("section", 2), ("room", 2)])
def test_pdf_scopes_produce_one_page_per_group(scope, pages):
    data = build_pdf(_sample_entries(), _SAMPLE_ROOMS, scope=scope, days=5)
    assert data.count(b"/Type /Page ") == pages


def test_pdf_handles_an_empty_timetable():
    assert build_pdf([], _SAMPLE_ROOMS, scope="all").startswith(b"%PDF")


def test_pdf_rejects_an_unknown_scope():
    with pytest.raises(ValueError):
        build_pdf(_sample_entries(), _SAMPLE_ROOMS, scope="galaxy")


def test_ics_is_well_formed_and_repeats_weekly():
    text = build_ics(_sample_entries(), start_date=date(2026, 1, 5), weeks=12)
    assert text.startswith("BEGIN:VCALENDAR\r\n") and text.rstrip().endswith("END:VCALENDAR")
    assert text.count("BEGIN:VEVENT") == 2 == text.count("END:VEVENT")
    assert "RRULE:FREQ=WEEKLY;COUNT=12" in text
    # Monday class must start on the Monday of that week, Wednesday two days later
    assert "DTSTART:20260105T083000" in text
    assert "DTSTART:20260107T133000" in text
    assert "\r\n" in text and all(len(line.encode()) <= 75 for line in text.split("\r\n"))


def test_ics_escapes_special_characters():
    entries = _sample_entries()
    entries[0]["course_name"] = "Maths; Stats, Advanced"
    assert r"Maths\; Stats\, Advanced" in build_ics(entries)


def test_filter_entries_narrows_by_teacher_section_and_room():
    entries = _sample_entries()
    assert len(filter_entries(entries, teacher="mr. saad salman")) == 1
    assert len(filter_entries(entries, course_id=102, section="b")) == 1
    assert len(filter_entries(entries, room_id=1)) == 1
    assert len(filter_entries(entries, shift="evening")) == 1
    assert len(filter_entries(entries)) == 2


def test_publish_endpoints_round_trip(client):
    assignments = [
        {"day": 1, "start_time": "08:30", "end_time": "09:50", "room_id": 1,
         "course_id": 101, "section": "A", "shift": "morning"},
        {"day": 3, "start_time": "10:00", "end_time": "11:20", "room_id": 2,
         "course_id": 102, "section": "A", "shift": "morning"},
    ]
    assert client.post("/api/timetable", json={"assignments": assignments}).status_code == 200

    targets = client.get("/api/publish/targets").get_json()
    assert targets["saved_classes"] == 2
    assert targets["teachers"] and targets["sections"] and targets["rooms"]

    pdf = client.post("/api/publish/pdf", json={"scope": "teacher", "days": 5})
    assert pdf.status_code == 200
    assert pdf.data.startswith(b"%PDF")
    assert pdf.mimetype == "application/pdf"

    ics = client.get("/calendar.ics?weeks=8")
    assert ics.status_code == 200
    assert ics.mimetype == "text/calendar"
    assert ics.data.count(b"BEGIN:VEVENT") == 2
    assert b"COUNT=8" in ics.data

    only_one = client.get(f"/calendar.ics?teacher={targets['teachers'][0]}")
    assert 1 <= only_one.data.count(b"BEGIN:VEVENT") < 3


def test_publish_uses_the_unsaved_grid_when_given_one(client):
    response = client.post("/api/publish/pdf", json={
        "scope": "all", "days": 1,
        "assignments": [{"day": 1, "start_time": "08:30", "end_time": "09:50", "room_id": 1,
                         "course_id": 101, "section": "A", "shift": "morning"}],
    })
    assert response.status_code == 200
    assert client.get("/api/timetable").get_json()["entries"] == []      # nothing was saved


def test_publish_rejects_an_empty_selection(client):
    assert client.post("/api/publish/pdf", json={"teacher": "Nobody At All"}).status_code == 400
    assert client.post("/api/publish/pdf", json={"scope": "moon"}).status_code == 400


# ------------------------------ building CRUD -------------------------------- #
def test_building_can_be_created_renamed_and_deleted(client):
    created = client.post("/api/buildings", json={"name": "Zeta Wing"})
    assert created.status_code == 201
    building_id = created.get_json()["id"]

    renamed = client.put(f"/api/buildings/{building_id}", json={"name": "Omega Wing"})
    assert renamed.status_code == 200 and renamed.get_json()["name"] == "Omega Wing"

    assert client.post("/api/buildings", json={"name": "omega wing"}).status_code == 400   # duplicate
    assert client.delete(f"/api/buildings/{building_id}").status_code == 200


def test_building_with_rooms_cannot_be_deleted(client):
    building_id = client.post("/api/buildings", json={"name": "Delta"}).get_json()["id"]
    client.post("/api/rooms", json={"room_number": "D-1", "building_name": "Delta", "capacity": 30})
    response = client.delete(f"/api/buildings/{building_id}")
    assert response.status_code == 400
    assert "room" in response.get_json()["message"].lower()


# --------------------------- front-end integrity ----------------------------- #
def _read(*parts) -> str:
    return (Path(__file__).resolve().parent.parent.joinpath(*parts)).read_text(encoding="utf-8")


def test_every_button_action_exists_in_the_controller():
    """A typo in a data-action would silently produce a dead button."""
    html = _read("timetable", "templates", "index.html")
    js = _read("timetable", "static", "app.js")
    actions = set(re.findall(r'data-action="([A-Za-z]+)"', html))
    declared = set(re.findall(r"^\s{4}([A-Za-z]+): ", js, re.M))
    assert actions, "the template should define some actions"
    assert actions <= declared, f"buttons with no handler: {sorted(actions - declared)}"


def test_every_shortcut_action_exists_in_the_controller():
    js = _read("timetable", "static", "app.js")
    used = set(re.findall(r'action: "([A-Za-z]+)"', js))
    declared = set(re.findall(r"^\s{4}([A-Za-z]+): ", js, re.M))
    assert used <= declared, f"shortcuts with no handler: {sorted(used - declared)}"


def test_every_dialog_referenced_by_the_controller_exists():
    html = _read("timetable", "templates", "index.html")
    js = _read("timetable", "static", "app.js")
    opened = set(re.findall(r'openDialog\("#([A-Za-z]+)"\)', js))
    present = set(re.findall(r'<div id="([A-Za-z]+)" class="dialog', html))
    assert opened <= present, f"missing dialogs: {sorted(opened - present)}"


def test_the_ui_no_longer_uses_blocking_browser_prompts():
    """Desktop webviews may ignore window.confirm/prompt - we ship our own."""
    js = "\n".join(
        line for line in _read("timetable", "static", "app.js").splitlines()
        if not line.lstrip().startswith(("//", "/*", "*"))
    )
    assert "window.confirm(" not in js
    assert js.count("window.prompt(") <= 1        # only the clipboard fallback


# ======================= v2.3: labs, semesters, gaps ========================= #
def test_catalogue_lists_a_separate_card_for_every_lab(client):
    catalogue = client.get("/api/courses").get_json()
    labs = [c for c in catalogue if c["kind"] == "lab"]
    assert labs, "the sample data should contain lab courses"
    for lab in labs:
        assert lab["has_lab"] is True
        assert lab["hours"] == lab["lab_credit_hours"] >= 1
        assert lab["key"].endswith(":lab")
        assert lab["label"].endswith("(Lab)")
        # the matching lecture exists too
        assert any(
            c["id"] == lab["id"] and c["section"] == lab["section"] and c["kind"] == "theory"
            for c in catalogue
        )


def test_a_lab_cannot_be_scheduled_for_a_course_without_one(client):
    theory = next(c for c in client.get("/api/courses").get_json() if not c["has_lab"])
    verdict = client.post("/api/timetable/validate", json={
        "candidate": {"day": 1, "start_time": "08:30", "end_time": "09:50", "room_id": 1,
                      "course_id": theory["id"], "section": theory["section"], "kind": "lab"},
        "grid": [],
    }).get_json()
    assert verdict["ok"] is False
    assert "no lab component" in verdict["conflicts"][0]["message"]


def test_a_lab_outside_a_lab_room_is_only_a_warning(client):
    lab = next(c for c in client.get("/api/courses").get_json() if c["kind"] == "lab")
    classroom = next(r for r in client.get("/api/rooms").get_json() if r["room_type"] != "Lab")
    verdict = client.post("/api/timetable/validate", json={
        "candidate": {"day": 2, "start_time": "08:30", "end_time": "09:50", "room_id": classroom["id"],
                      "course_id": lab["id"], "section": lab["section"], "kind": "lab"},
        "grid": [],
    }).get_json()
    assert verdict["ok"] is True                       # a warning never blocks
    assert [c["kind"] for c in verdict["conflicts"]] == ["roomtype"]
    assert verdict["conflicts"][0]["severity"] == "warning"


def test_lecture_and_lab_of_one_section_cannot_overlap(client):
    lab = next(c for c in client.get("/api/courses").get_json() if c["kind"] == "lab")
    slot = {"day": 3, "start_time": "10:00", "end_time": "11:20",
            "course_id": lab["id"], "section": lab["section"]}
    verdict = client.post("/api/timetable/validate", json={
        "candidate": dict(slot, room_id=2, kind="lab"),
        "grid": [dict(slot, room_id=1, kind="theory")],
    }).get_json()
    assert verdict["ok"] is False
    duplicate = next(c for c in verdict["conflicts"] if c["kind"] == "duplicate")
    assert "same students attend both" in duplicate["message"]


def test_two_courses_of_the_same_semester_and_section_clash(client):
    catalogue = client.get("/api/courses").get_json()
    first = next(c for c in catalogue if c["semester"] and c["kind"] == "theory")
    second = next(
        c for c in catalogue
        if c["kind"] == "theory" and c["semester"] == first["semester"]
        and c["section"] == first["section"] and c["id"] != first["id"]
    )
    verdict = client.post("/api/timetable/validate", json={
        "candidate": {"day": 4, "start_time": "08:30", "end_time": "09:50", "room_id": 2,
                      "course_id": second["id"], "section": second["section"]},
        "grid": [{"day": 4, "start_time": "08:30", "end_time": "09:50", "room_id": 1,
                  "course_id": first["id"], "section": first["section"]}],
    }).get_json()
    assert verdict["ok"] is False
    kinds = [c["kind"] for c in verdict["conflicts"]]
    assert "semester" in kinds
    message = next(c for c in verdict["conflicts"] if c["kind"] == "semester")["message"]
    assert f"Semester {first['semester']} section {first['section']}" in message


def test_unscheduled_report_lists_missing_lectures_and_labs(client):
    everything = client.post("/api/timetable/unscheduled", json={"assignments": []}).get_json()
    assert everything["ok"] is False
    assert everything["required"] == 65 == len(everything["missing"])
    assert sum(everything["by_semester"].values()) == 65
    assert {item["kind"] for item in everything["missing"]} == {"theory", "lab"}

    created = client.post("/api/timetable/autofill", json={
        "assignments": [], "days": 5,
        "slots": [{"start": f"{8 + i * 2:02d}:30", "end": f"{9 + i * 2:02d}:50"} for i in range(4)],
        "room_ids": [r["id"] for r in client.get("/api/rooms").get_json()][:14],
    }).get_json()["created"]
    filled = client.post("/api/timetable/unscheduled", json={"assignments": created}).get_json()
    assert filled == {"ok": True, "required": 65, "missing": [], "by_semester": {}}

    # ... and the saved timetable is used when no grid is supplied
    client.post("/api/timetable", json={"assignments": created[:-2]})
    saved = client.post("/api/timetable/unscheduled", json={}).get_json()
    assert len(saved["missing"]) == 2


def test_autofill_can_be_limited_to_one_semester(client):
    created = client.post("/api/timetable/autofill", json={
        "assignments": [], "days": 5, "semester": 3,
        "slots": [{"start": "08:30", "end": "09:50"}, {"start": "10:00", "end": "11:20"}],
        "room_ids": [1, 2, 3, 4, 5, 6],
    }).get_json()["created"]
    assert created
    semesters = {
        c["semester"] for c in client.get("/api/courses").get_json()
        if any(c["id"] == entry["course_id"] for entry in created)
    }
    assert semesters == {3}


def test_publish_pdf_accepts_a_semester_scope(client):
    client.post("/api/timetable", json={"assignments": [
        {"day": 1, "start_time": "08:30", "end_time": "09:50", "room_id": 1,
         "course_id": 101, "section": "A"},
    ]})
    targets = client.get("/api/publish/targets").get_json()
    assert targets["semesters"]
    assert targets["unscheduled"] == 64

    response = client.post("/api/publish/pdf", json={"scope": "semester", "days": 5})
    assert response.status_code == 200
    assert response.data.startswith(b"%PDF-1.4")

    bad = client.post("/api/publish/pdf", json={"scope": "nonsense"})
    assert bad.status_code == 400


def test_course_editor_round_trips_the_lab_and_semester_fields(client):
    created = client.post("/api/courses", json={
        "code": "SE4001", "name": "Compiler Construction", "credit_hours": 3,
        "semester": 6, "has_lab": "yes", "lab_credit_hours": 2,
        "sections": [{"section": "A"}],
    }).get_json()
    assert created["semester"] == 6 and created["has_lab"] == 1 and created["lab_credit_hours"] == 2

    catalogue = client.get("/api/courses").get_json()
    mine = [c for c in catalogue if c["id"] == created["id"]]
    assert sorted(c["kind"] for c in mine) == ["lab", "theory"]

    # turning the lab off clears its credit hours and removes the lab card
    updated = client.put(f"/api/courses/{created['id']}", json={
        "code": "SE4001", "name": "Compiler Construction", "semester": 6, "has_lab": "",
        "sections": [{"section": "A"}],
    }).get_json()
    assert updated["has_lab"] == 0 and updated["lab_credit_hours"] == 0
    assert [c["kind"] for c in client.get("/api/courses").get_json() if c["id"] == created["id"]] == ["theory"]


def test_import_template_carries_the_lab_and_semester_columns(client):
    response = client.get("/api/import/template")
    if response.status_code == 501:
        pytest.skip("openpyxl is not installed")
    import io as _io

    import openpyxl

    workbook = openpyxl.load_workbook(_io.BytesIO(response.data))
    headers = [cell.value for cell in workbook["Courses"][1]]
    assert "Semester (1-12, blank = none)" in headers
    assert "Has lab? (yes/no)" in headers
    assert "Lab credit hours" in headers


def test_the_shortcut_hints_are_no_longer_printed_on_the_controls():
    """Alt+1 and friends belong in the F1 reference, not on every button."""
    html = _read("timetable", "templates", "index.html")
    controls = re.findall(r"<button[^>]*>.*?</button>", html, re.S)
    assert controls
    assert not [button for button in controls if "<kbd>" in button]
    # but the reference itself still lists them
    assert html.count("<kbd>") > 10
    assert 'title="Morning shift (Alt+1)"' in html


# ======================= v2.4: Class Schedule printed list ===================== #


@pytest.mark.parametrize("start,end,expected", [
    ("14:30", "16:00", "2:30 PM - 4:00 PM"),
    ("08:30", "09:50", "8:30 AM - 9:50 AM"),
    ("00:00", "12:00", "12:00 AM - 12:00 PM"),
    ("23:00", "23:30", "11:00 PM - 11:30 PM"),
])
def test_format_time_range_uses_am_pm(start, end, expected):
    assert format_time_range(start, end) == expected


def _schedule_entries():
    """A tiny timetable including a zero-credit (Non-credited) course."""
    return [
        {
            "id": 1, "day": 1, "start_time": "14:30", "end_time": "16:00", "shift": "morning",
            "room_id": 1, "room_number": "105", "room_label": "A-105", "course_id": 101,
            "code": "CHEM4134", "course_name": "Special Paper - I", "section": "A",
            "color": "#a9d2e1", "instructor": "Miss Fozia", "num_students": 42,
            "semester": 4, "kind": "theory", "credit_hours": 3,
        },
        {
            "id": 2, "day": 1, "start_time": "16:00", "end_time": "17:30", "shift": "morning",
            "room_id": 1, "room_number": "105", "room_label": "A-105", "course_id": 102,
            "code": "CHEM4130", "course_name": "Food and Drug Analysis", "section": "A",
            "color": "#a9d2e1", "instructor": "Mr Umar Faraz", "num_students": 40,
            "semester": 4, "kind": "theory", "credit_hours": 3,
        },
        {
            "id": 3, "day": 4, "start_time": "17:30", "end_time": "18:30", "shift": "morning",
            "room_id": 9, "room_number": "mosque", "room_label": "mosque", "course_id": 120,
            "code": "", "course_name": "Tajuma Quran course", "section": "A",
            "color": "#a9d2e1", "instructor": "Dr. Bilal", "num_students": 0,
            "semester": 4, "kind": "theory", "credit_hours": 0,
        },
    ]


_SCHEDULE_ROOMS = [
    {"id": 1, "room_number": "105", "label": "A-105", "capacity": 60, "room_type": "Classroom"},
    {"id": 9, "room_number": "mosque", "label": "mosque", "capacity": 0, "room_type": "Classroom"},
]


@pytest.mark.skipif(not openpyxl_available(), reason="openpyxl not installed")
def test_class_schedule_layout_matches_the_reference():
    import io as _io

    from openpyxl import load_workbook

    data = build_workbook(
        _schedule_entries(), _SCHEDULE_ROOMS, days=5, layout="schedule",
        title="Class Schedule", term="Spring 2024",
        institution="University of Education Jauharabad Campus",
        program="BS Chemistry (Post ADP)", semester="4", commencement="January 2024",
    )
    book = load_workbook(_io.BytesIO(data))
    assert book.sheetnames == ["Class Schedule"]
    sheet = book["Class Schedule"]

    # metadata title block
    assert sheet.cell(row=1, column=1).value == "Class Schedule Spring 2024"
    assert sheet.cell(row=2, column=1).value == "University of Education Jauharabad Campus"
    assert "Name of program: BS Chemistry (Post ADP)" in str(sheet.cell(row=3, column=1).value)
    assert "Semester: 4" in str(sheet.cell(row=3, column=5).value)
    assert sheet.cell(row=4, column=1).value == "Commencement of Classes: January 2024"

    # header row (find it after the metadata block)
    header = None
    for r in range(1, 8):
        if sheet.cell(row=r, column=1).value == "Days":
            header = r
            break
    assert header is not None, "header row not found"
    expected = ["Days", "Course Code", "Course Title", "C.Hrs", "Total No.of Students",
                "Teacher's Name", "Time", "Room No"]
    assert [sheet.cell(row=header, column=c).value for c in range(1, 9)] == expected

    # Monday block
    mon = header + 1
    assert sheet.cell(row=mon, column=1).value == "Monday"
    assert sheet.cell(row=mon, column=2).value == "CHEM4134"
    assert sheet.cell(row=mon, column=4).value == "3"           # credit hours
    assert sheet.cell(row=mon, column=7).value == "2:30 PM - 4:00 PM"   # AM/PM 12h range
    assert sheet.cell(row=mon, column=8).value == "105"

    # Non-credited course (credit hours == 0)
    thu = next(r for r in range(header + 1, sheet.max_row + 1)
               if sheet.cell(row=r, column=3).value == "Tajuma Quran course")
    assert sheet.cell(row=thu, column=1).value == "Thursday"
    assert sheet.cell(row=thu, column=4).value == "non-credited course"


@pytest.mark.skipif(not openpyxl_available(), reason="openpyxl not installed")
def test_book_layout_is_the_default_when_layout_is_omitted():
    import io as _io

    from openpyxl import load_workbook

    data = build_workbook(_schedule_entries(), _SCHEDULE_ROOMS, days=1)
    book = load_workbook(_io.BytesIO(data))
    assert "Class Schedule" not in book.sheetnames      # that is layout="schedule"
    assert "Summary" in book.sheetnames and "Monday" in book.sheetnames
    assert book.sheetnames[0] == "Contents"


def test_grid_layout_still_draws_room_by_time_grids():
    """layout="grid" keeps the facilities view: rooms down the side, slots across."""
    import io as _io

    from openpyxl import load_workbook

    data = build_workbook(_schedule_entries(), _SCHEDULE_ROOMS, days=1, layout="grid")
    book = load_workbook(_io.BytesIO(data))
    assert "Contents" in book.sheetnames
    monday = book["Monday"]
    assert monday.cell(row=4, column=1).value == "Room / Time"
    assert "A-105" in str(monday.cell(row=5, column=1).value)
    # the grid cells carry the course label and the teacher, not a reference row
    assert "Special Paper - I" in str(monday.cell(row=5, column=2).value)


def test_credit_hours_are_carried_into_export_entries(service):
    """describe_assignments must expose credit_hours so exports can spot non-credited courses."""
    from timetable.services import Assignment

    entry = service.describe_assignments([
        Assignment(day=1, start_time="08:30", end_time="09:50", room_id=1,
                   course_id=101, section="A")
    ])[0]
    assert "credit_hours" in entry
    assert entry["credit_hours"] >= 1


def test_schedule_pdf_renders_a_single_landscape_page():
    """layout='schedule' on build_pdf must produce a valid single-page document."""
    from timetable.publishing import build_pdf

    data = build_pdf(
        _schedule_entries(), _SCHEDULE_ROOMS, days=5, layout="schedule",
        title="Class Schedule", term="Spring 2024",
        institution="University of Education Jauharabad Campus",
        program="BS Chemistry (Post ADP)", semester="4", commencement="January 2024",
    )
    assert data.startswith(b"%PDF-1.4")
    assert data.count(b"/Type /Page ") == 1


# =================== v2.5: standalone desktop (no server) ==================== #
def test_standalone_html_is_self_contained(app):
    """The native window must not depend on a Flask static server at all."""
    from timetable.desktop import _build_standalone_html

    html = _build_standalone_html(app)
    # assets are inlined, not referenced from /static
    assert "rel=\"stylesheet\" href=\"/static" not in html
    assert "<script defer src=\"/static" not in html
    assert html.count("<style>") >= 1
    assert "front-end controller" in html            # app.js is embedded
    # the in-process fetch shim is injected before the controller
    assert "window.__NATIVE__" in html
    assert "window.pywebview.api" in html
    # renders the title/version without error
    assert "Automated Timetable Generator" in html


def test_native_bridge_routes_requests_in_process(app):
    """The desktop bridge must answer the same endpoints as the HTTP server."""
    from timetable.desktop import _NativeBridge

    bridge = _NativeBridge(app)

    health = bridge.request({"method": "GET", "path": "/api/health", "query": ""})
    assert health["status"] == 200
    assert '"status":"ok"' in health["text"]

    validate = bridge.request({
        "method": "POST", "path": "/api/timetable/validate", "query": "", "raw": False,
        "body": '{"candidate":{"day":1,"start_time":"08:30","end_time":"09:50",'
                '"room_id":1,"course_id":101,"section":"A"},"grid":[]}',
    })
    assert validate["status"] == 200
    assert '"ok":true' in validate["text"]

    wrong = bridge.request({"method": "GET", "path": "/api/does-not-exist", "query": ""})
    assert wrong["status"] == 404


def test_native_bridge_streams_binary_exports(app):
    """Raw (binary) exports come back as base64 so the window can save them."""
    import base64

    from timetable.desktop import _NativeBridge

    bridge = _NativeBridge(app)
    body = '{"assignments":[{"day":1,"start_time":"08:30","end_time":"09:50",' \
           '"room_id":1,"course_id":101,"section":"A","shift":"morning"}],' \
           '"days":2,"shift":"all"}'
    result = bridge.request({
        "method": "POST", "path": "/api/export/csv", "query": "", "raw": True, "body": body,
    })
    assert result["status"] == 200
    raw = base64.b64decode(result["base64"])
    assert raw.startswith("\ufeff".encode("utf-8"))     # Excel-friendly BOM
    assert b"Monday" in raw


# ======================= v2.6: reports, day scope, autosave ==================== #
from timetable.reports import conflict_report, room_utilisation, teacher_workload  # noqa: E402
from timetable.exporters import build_workbook  # noqa: E402


def _report_entries():
    """Three classes with one room conflict and a 0-credit course."""
    return [
        {
            "id": 1, "day": 1, "start_time": "08:30", "end_time": "09:50", "shift": "morning",
            "room_id": 1, "room_number": "101", "room_label": "A-101", "course_id": 101,
            "code": "CS3009", "course_name": "Artificial Intelligence", "section": "A",
            "kind": "theory", "semester": 1, "instructor": "Dr A", "num_students": 40,
            "credit_hours": 3, "capacity": 60,
        },
        {
            "id": 2, "day": 1, "start_time": "08:30", "end_time": "09:50", "shift": "morning",
            "room_id": 1, "room_number": "101", "room_label": "A-101", "course_id": 102,
            "code": "CS4001", "course_name": "Machine Learning", "section": "B",
            "kind": "theory", "semester": 2, "instructor": "Dr B", "num_students": 30,
            "credit_hours": 3, "capacity": 60,
        },
        {
            "id": 3, "day": 2, "start_time": "17:30", "end_time": "18:30", "shift": "morning",
            "room_id": 9, "room_number": "mosque", "room_label": "mosque", "course_id": 120,
            "code": "", "course_name": "Tajuma Quran course", "section": "A",
            "kind": "theory", "semester": 4, "instructor": "Dr Bilal", "num_students": 0,
            "credit_hours": 0, "capacity": 0,
        },
    ]


_REPORT_ROOMS = [
    {"id": 1, "room_number": "101", "label": "A-101", "capacity": 60, "room_type": "Classroom"},
    {"id": 9, "room_number": "mosque", "label": "mosque", "capacity": 0, "room_type": "Classroom"},
]


def test_room_utilisation_report_flags_under_used_rooms():
    slots = [{"start": "08:30", "end": "09:50"}, {"start": "10:00", "end": "11:20"}]
    report = room_utilisation(_report_entries(), _REPORT_ROOMS, days=5, slots=slots)
    assert report["headers"][0] == "Room"
    assert "Utilisation" in report["headers"]

    a101 = next(r for r in report["rows"] if r[0] == "A-101")
    assert a101[3] == 2                       # two classes per week
    assert a101[6].endswith("%")
    # the 0-credit course is in the "mosque" room of capacity 0 -> under-used
    mosque = next(r for r in report["rows"] if r[0] == "mosque")
    assert mosque[6].endswith("%")


def test_teacher_workload_report_orders_by_hours():
    report = teacher_workload(_report_entries())
    assert "Contact hours / week" in report["headers"]
    hours = [row[2] for row in report["rows"]]
    assert hours == sorted(hours, reverse=True)
    dr_a = next(r for r in report["rows"] if r[0] == "Dr A")
    assert dr_a[1] == 1                       # one theory class


def test_conflict_report_sorts_errors_before_warnings():
    report = conflict_report(_report_entries())
    assert report["rows"]
    severities = [row[0] for row in report["rows"]]
    # All "Error" rows come before any "Warning" row.
    try:
        first_warning = severities.index("Warning")
    except ValueError:
        first_warning = len(severities)
    assert all(severity == "Error" for severity in severities[:first_warning])
    assert any("Same room booked twice" in row[7] for row in report["rows"])


@pytest.mark.parametrize("scope,pages", [
    ("day", 2),          # two days with classes -> two pages
])
def test_publish_day_scope_is_one_page_per_day(scope, pages):
    from timetable.publishing import build_pdf
    data = build_pdf(_report_entries(), _REPORT_ROOMS, days=5, scope=scope, layout="grid")
    assert data.startswith(b"%PDF-1.4")
    assert data.count(b"/Type /Page ") == pages


@pytest.mark.parametrize("scope", ["utilisation", "workload", "conflict"])
def test_publish_report_scopes_produce_valid_pdfs(scope):
    from timetable.publishing import build_pdf
    data = build_pdf(_report_entries(), _REPORT_ROOMS, days=5, scope=scope, layout="grid")
    assert data.startswith(b"%PDF-1.4")
    assert data.count(b"/Type /Page ") == 1


def test_workbook_contains_the_report_sheets():
    import io as _io
    from openpyxl import load_workbook

    for layout in ("grid", "book"):
        data = build_workbook(_report_entries(), _REPORT_ROOMS, days=5, layout=layout)
        book = load_workbook(_io.BytesIO(data))
        assert "Room Utilisation" in book.sheetnames
        assert "Teacher Workload" in book.sheetnames
        assert "Conflict Report" in book.sheetnames
        sheet = book["Conflict Report"]
        header = _header_row(sheet, "Severity")
        assert sheet.cell(row=header, column=1).value == "Severity"
        assert sheet["A1"].value == "Conflict Report"


def test_project_autosave_writes_a_backup_next_to_the_project(app, tmp_path):
    from timetable.projects import autosave_project

    client = app.test_client()
    project = client.post("/api/project/save", json={
        "name": "Autosave Test", "path": str(tmp_path / "proj")
    }).get_json()
    assert project["ok"] is True

    engine = app.extensions["engine"]
    backup = autosave_project(engine, project["path"], "Autosave Test")
    assert backup["ok"] is True
    backup_path = backup["path"]
    assert backup_path.endswith(".ttproj")
    parent = Path(backup_path).parent
    assert parent.name == "_backups"
    assert Path(backup_path).is_file()
    assert len(list(parent.glob("proj-*.ttproj"))) >= 1


def test_project_autosave_via_api(app, tmp_path):
    client = app.test_client()
    project = client.post("/api/project/save", json={
        "name": "API Autosave", "path": str(tmp_path / "api-proj")
    }).get_json()
    assert project["ok"] is True

    result = client.post("/api/project/autosave", json={"path": project["path"]}).get_json()
    assert result["ok"] is True
    assert Path(result["path"]).is_file()


# ----------------------- v2.6: /api/report endpoint -------------------------- #
def test_api_report_room_utilisation(client):
    assignments = [
        {"day": 1, "start_time": "08:30", "end_time": "09:50", "room_id": 1,
         "course_id": 101, "section": "A", "shift": "morning"},
        {"day": 1, "start_time": "10:00", "end_time": "11:20", "room_id": 1,
         "course_id": 102, "section": "A", "shift": "morning"},
    ]
    client.post("/api/timetable", json={"assignments": assignments})
    report = client.post("/api/report/utilisation", json={"days": 5}).get_json()
    assert report["ok"] is True
    assert report["title"]
    assert report["headers"][0] == "Room"
    assert report["entries"] == 2
    assert any("Utilisation" in str(h) for h in report["headers"])


def test_api_report_teacher_workload(client):
    client.post("/api/timetable", json={"assignments": [
        {"day": 1, "start_time": "08:30", "end_time": "09:50", "room_id": 1,
         "course_id": 101, "section": "A", "shift": "morning"},
    ]})
    report = client.post("/api/report/workload", json={"days": 5}).get_json()
    assert report["ok"] is True
    assert report["headers"][0] == "Teacher"
    assert report["rows"]


def test_api_report_conflict(client):
    assignments = [
        {"day": 1, "start_time": "08:30", "end_time": "09:50", "room_id": 1,
         "course_id": 101, "section": "A", "shift": "morning"},
        {"day": 1, "start_time": "08:30", "end_time": "09:50", "room_id": 1,
         "course_id": 102, "section": "A", "shift": "morning"},
    ]
    # These two genuinely clash, so the report must see both of them: pass the
    # on-screen grid straight in rather than going through the (blocking) save.
    report = client.post("/api/report/conflict", json={"assignments": assignments, "days": 5}).get_json()
    assert report["ok"] is True
    assert report["rows"]
    assert report["headers"][0] == "Severity"
    assert any("Same room booked twice" in str(row) for row in report["rows"])


def test_api_report_rejects_an_unknown_scope(client):
    response = client.post("/api/report/nonsense", json={})
    assert response.status_code == 400


@pytest.mark.skipif(not openpyxl_available(), reason="openpyxl not installed")
def test_report_sheets_appear_in_the_workbook_api(client):
    client.post("/api/timetable", json={"assignments": [
        {"day": 1, "start_time": "08:30", "end_time": "09:50", "room_id": 1,
         "course_id": 101, "section": "A", "shift": "morning"},
    ]})
    from openpyxl import load_workbook
    import io as _io

    response = client.post("/api/export/xlsx", json={"days": 5, "shift": "all"})
    book = load_workbook(_io.BytesIO(response.data))
    assert "Room Utilisation" in book.sheetnames
    assert "Teacher Workload" in book.sheetnames
    assert "Conflict Report" in book.sheetnames
    conflict = book["Conflict Report"]
    # Even with a valid timetable the Conflict Report sheet must exist.
    assert conflict["A1"].value == "Conflict Report"

    # the daily scope publishes one landscape page per weekday via the API too
    day_response = client.post("/api/publish/pdf", json={"scope": "day", "days": 5})
    assert day_response.status_code == 200
    assert day_response.data.startswith(b"%PDF-1.4")
    assert day_response.mimetype == "application/pdf"


def test_report_and_day_scopes_win_over_schedule_layout():
    """layout='schedule' alone must not swallow the report / per-day scopes."""
    from timetable.publishing import build_pdf
    data = build_pdf(_report_entries(), _REPORT_ROOMS, days=5, scope="utilisation", layout="schedule")
    assert data.startswith(b"%PDF-1.4")
    assert data.count(b"/Type /Page ") == 1
    day_data = build_pdf(_report_entries(), _REPORT_ROOMS, days=5, scope="day", layout="schedule")
    assert day_data.count(b"/Type /Page ") == 2


def test_file_in_use_message_is_friendly():
    from timetable.web import _file_in_use_message
    from pathlib import Path
    import errno as _errno
    target = Path("/tmp/some-workbook.xlsx")
    msg = _file_in_use_message(target, OSError(_errno.EBUSY, "busy"))
    assert "open in another program" in msg
    assert target.name in msg
    assert "nothing was changed" in msg
    generic = _file_in_use_message(target, OSError(_errno.EACCES, "denied"))
    assert generic


# ==================== v2.1.0: the semester-book Excel export ================== #


def _book(**overrides):
    """Build the default (book) workbook from the shared reference fixture."""
    import io as _io

    from openpyxl import load_workbook

    kwargs = dict(
        days=5, title="Class Schedule", term="Spring 2024",
        institution="University of Education Jauharabad Campus",
        program="BS Chemistry (Post ADP)", semester="4", commencement="January 2024",
        unscheduled=[{"course_id": 199, "code": "CHEM4199", "course_name": "Project Work",
                      "section": "A", "kind": "theory", "semester": 4, "hours": 3,
                      "instructor": "Dr Project"}],
    )
    kwargs.update(overrides)
    layout = kwargs.pop("layout", "book")
    data = build_workbook(_schedule_entries(), _SCHEDULE_ROOMS, layout=layout, **kwargs)
    return load_workbook(_io.BytesIO(data))


@pytest.mark.skipif(not openpyxl_available(), reason="openpyxl not installed")
def test_semester_book_opens_with_a_hyperlinked_contents_page():
    book = _book()
    assert book.sheetnames[0] == "Contents"
    contents = book["Contents"]

    listed = [contents.cell(row=r, column=1).value for r in range(1, contents.max_row + 1)]
    for name in book.sheetnames[1:]:
        assert name in listed, f"{name} is missing from the Contents page"

    # the sheet names are real links into the workbook, not just text
    linked = [
        cell for row in contents.iter_rows() for cell in row
        if cell.hyperlink is not None and cell.hyperlink.location
    ]
    assert linked, "the Contents page carries no hyperlinks"
    targets = {cell.hyperlink.location for cell in linked}
    assert "'Summary'!A1" in targets
    assert "'Semester 4'!A1" in targets


@pytest.mark.skipif(not openpyxl_available(), reason="openpyxl not installed")
def test_semester_sheet_uses_the_reference_arrangement_and_totals():
    book = _book()
    # the fixture only has semester-4 classes, so there is exactly one sheet
    assert [n for n in book.sheetnames if n.startswith("Semester ")] == ["Semester 4"]
    sheet = book["Semester 4"]
    header = _header_row(sheet)
    assert [sheet.cell(row=header, column=c).value for c in range(1, 9)] == [
        "Days", "Course Code", "Course Title", "C.Hrs", "Total No.of Students",
        "Teacher's Name", "Time", "Room No",
    ]

    # metadata title block, exactly like the printed Class Schedule
    assert sheet.cell(row=1, column=1).value == "Class Schedule Spring 2024 — Semester 4"
    assert sheet.cell(row=2, column=1).value == "University of Education Jauharabad Campus"
    assert "Name of program: BS Chemistry (Post ADP)" in str(sheet.cell(row=3, column=1).value)
    assert sheet.cell(row=4, column=1).value == "Commencement of Classes: January 2024"

    # Monday block: two classes, day cell merged over them
    monday = header + 1
    assert sheet.cell(row=monday, column=1).value == "Monday"
    assert sheet.cell(row=monday, column=2).value == "CHEM4134"
    assert sheet.cell(row=monday, column=7).value == "2:30 PM - 4:00 PM"
    assert sheet.cell(row=monday + 1, column=1).value is None       # merged away
    assert any("CHEM4130" == sheet.cell(row=r, column=2).value for r in range(monday, monday + 3))

    # the non-credited course keeps its label, and the totals strip is present
    texts = [str(cell.value) for row in sheet.iter_rows() for cell in row if cell.value]
    assert any("non-credited course" == text for text in texts)
    assert any(text.startswith("Total: 3 classes") for text in texts)


@pytest.mark.skipif(not openpyxl_available(), reason="openpyxl not installed")
def test_weekday_sheets_group_the_day_by_semester_and_section():
    book = _book()
    monday = book["Monday"]
    header = _header_row(monday, "Section")
    assert monday.cell(row=header, column=1).value == "Section"
    assert monday.cell(row=header + 1, column=1).value == "Semester 4\nSection A"
    assert monday.cell(row=header + 1, column=2).value == "CHEM4134"

    # an empty weekday still gets its sheet, and says so instead of lying
    tuesday = book["Tuesday"]
    tuesday_header = _header_row(tuesday, "Section")
    assert "No classes are scheduled on Tuesday." == tuesday.cell(
        row=tuesday_header + 1, column=1).value


@pytest.mark.skipif(not openpyxl_available(), reason="openpyxl not installed")
def test_by_teacher_sheet_merges_the_teacher_over_their_classes():
    book = _book()
    sheet = book["By Teacher"]
    header = _header_row(sheet, "Teacher")
    assert [sheet.cell(row=header, column=c).value for c in range(1, 9)] == [
        "Teacher", "Day", "Course Code", "Course Title", "C.Hrs",
        "Total No.of Students", "Time", "Room No",
    ]
    teachers = [sheet.cell(row=r, column=1).value for r in range(header + 1, sheet.max_row + 1)]
    assert "Miss Fozia" in teachers and "Dr. Bilal" in teachers


@pytest.mark.skipif(not openpyxl_available(), reason="openpyxl not installed")
def test_credit_hour_audit_reports_planned_versus_scheduled_hours():
    book = _book()
    sheet = book["Credit Hour Audit"]
    header = _header_row(sheet, "Semester")
    assert sheet.cell(row=header, column=7).value == "Planned hrs / week"
    assert sheet.cell(row=header, column=8).value == "Contact hrs on grid"

    rows = {}
    for r in range(header + 1, sheet.max_row + 1):
        code = sheet.cell(row=r, column=2).value
        if code:
            rows[code] = [sheet.cell(row=r, column=c).value for c in range(7, 11)]

    # CHEM4134 is a 3-credit course booked for one 90-minute period a week
    assert rows["CHEM4134"][0] == 3
    assert rows["CHEM4134"][1] == 1.5
    assert str(rows["CHEM4134"][3]).startswith("Short")
    # CHEM4199 was never placed on the grid at all
    assert rows["CHEM4199"][3] == "Not scheduled"


@pytest.mark.skipif(not openpyxl_available(), reason="openpyxl not installed")
def test_dashboard_sheet_carries_kpis_and_charts():
    book = _book()
    sheet = book["Dashboard"]
    header = _header_row(sheet, "Measure")
    measures = [sheet.cell(row=r, column=1).value for r in range(header + 1, sheet.max_row + 1)]
    assert "Scheduled classes / week" in measures
    assert "Classes still unscheduled" in measures
    assert sheet.cell(row=header + 1, column=2).value == 3          # three scheduled classes
    assert len(sheet._charts) == 2                                  # utilisation + workload
    # the chart source columns are hidden, so the charts must be allowed to
    # read them or Excel draws two empty frames
    assert all(chart.visible_cells_only is False for chart in sheet._charts)
    assert sheet.column_dimensions["H"].hidden is True


@pytest.mark.skipif(not openpyxl_available(), reason="openpyxl not installed")
def test_master_data_sheet_lists_courses_teachers_and_rooms():
    book = _book()
    sheet = book["Master Data"]
    titles = [str(cell.value) for row in sheet.iter_rows() for cell in row]
    assert "Courses" in titles and "Teachers" in titles and "Rooms" in titles
    assert "CHEM4134" in titles
    assert "Miss Fozia" in titles
    assert "mosque" in titles


@pytest.mark.skipif(not openpyxl_available(), reason="openpyxl not installed")
def test_the_extra_sheets_can_be_switched_off():
    book = _book(show_audit=False, show_dashboard=False, show_master_data=False,
                 show_unscheduled=False, contents=False)
    for absent in ("Credit Hour Audit", "Dashboard", "Master Data", "Unscheduled", "Contents"):
        assert absent not in book.sheetnames
    # the core document survives
    assert "Semester 4" in book.sheetnames and "Summary" in book.sheetnames
    assert "Room Utilisation" in book.sheetnames                    # reports always travel


@pytest.mark.skipif(not openpyxl_available(), reason="openpyxl not installed")
def test_every_cell_uses_the_chosen_font_and_prints_one_page_wide():
    book = _book(font_name="Georgia", font_size=11)
    for sheet in book.worksheets:
        assert sheet.page_setup.fitToWidth == 1
        if sheet.sheet_properties.tabColor:
            assert str(sheet.sheet_properties.tabColor.rgb).startswith("FF")   # opaque tab
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    assert cell.font.name == "Georgia", (sheet.title, cell.coordinate)


@pytest.mark.skipif(not openpyxl_available(), reason="openpyxl not installed")
def test_grid_layout_is_untouched_by_the_book_only_sheets():
    book = _book(layout="grid")
    assert "Contents" in book.sheetnames and "Monday" in book.sheetnames
    for absent in ("Credit Hour Audit", "Dashboard", "Master Data"):
        assert absent not in book.sheetnames
    assert book["Monday"].cell(row=4, column=1).value == "Room / Time"


@pytest.mark.skipif(not openpyxl_available(), reason="openpyxl not installed")
def test_csv_bundle_has_one_file_per_sheet(tmp_path):
    import io as _io
    import zipfile

    from timetable.exporters import build_csv_bundle

    blob = build_csv_bundle(
        _schedule_entries(),
        unscheduled=[{"course_id": 199, "code": "CHEM4199", "course_name": "Project Work",
                      "section": "A", "kind": "theory", "semester": 4, "hours": 3,
                      "instructor": "Dr Project"}],
    )
    archive = zipfile.ZipFile(_io.BytesIO(blob))
    names = set(archive.namelist())
    assert "timetable.csv" in names
    assert "semester-4.csv" in names
    assert "monday.csv" in names and "thursday.csv" in names
    assert "tuesday.csv" not in names                 # nothing happens on Tuesday
    assert "by-teacher.csv" in names
    assert "credit-hour-audit.csv" in names
    assert "unscheduled.csv" in names

    head = archive.read("semester-4.csv").decode("utf-8")
    assert head.startswith("\ufeff")                  # Excel-friendly BOM
    assert '"CHEM4134"' in head


@pytest.mark.skipif(not openpyxl_available(), reason="openpyxl not installed")
def test_csv_bundle_endpoint_writes_a_zip_next_to_the_project(client, tmp_path):
    client.post("/api/timetable", json={"assignments": [
        {"day": 1, "start_time": "08:30", "end_time": "09:50", "room_id": 1,
         "course_id": 101, "section": "A", "shift": "morning"},
    ]})
    response = client.post("/api/export/csv-bundle", json={"folder": str(tmp_path)})
    assert response.status_code == 200
    payload = response.get_json()
    assert payload["saved"] is True

    import zipfile

    written = Path(payload["path"])
    assert written.suffix == ".zip"
    assert "timetable.csv" in zipfile.ZipFile(written).namelist()


@pytest.mark.skipif(not openpyxl_available(), reason="openpyxl not installed")
def test_saved_export_settings_are_honoured_whichever_case_they_use(client):
    """The dialog persists camelCase keys; the exporters take snake_case.  Both
    must reach the workbook, otherwise a saved font choice silently never
    applies."""
    client.post("/api/timetable", json={"assignments": [
        {"day": 1, "start_time": "08:30", "end_time": "09:50", "room_id": 1,
         "course_id": 101, "section": "A", "shift": "morning"},
    ]})
    client.post("/api/settings", json={"export": {
        "fontName": "Georgia", "showDashboard": False, "institution": "Saved Institute",
    }})

    import io as _io

    from openpyxl import load_workbook

    data = client.post("/api/export/xlsx", json={}).data
    book = load_workbook(_io.BytesIO(data))
    assert "Dashboard" not in book.sheetnames
    semester = book["Semester 1"]
    assert semester["A2"].value == "Saved Institute"
    for row in semester.iter_rows():
        for cell in row:
            if cell.value is not None:
                assert cell.font.name == "Georgia"

    # an explicit snake_case request still wins over the saved setting
    data = client.post("/api/export/xlsx", json={"font_name": "Arial", "show_dashboard": True}).data
    book = load_workbook(_io.BytesIO(data))
    assert "Dashboard" in book.sheetnames
    assert book["Summary"]["A1"].font.name == "Arial"


def test_export_prose_is_pluralised_not_parenthesised():
    """The workbook is a document people read, so it must say "1 class" and
    "3 classes" - never "1 class(es)" and never "3 classs"."""
    from timetable.exporters import _plural, _word

    assert _plural(1, "class") == "1 class"
    assert _plural(3, "class") == "3 classes"          # not "classs"
    assert _plural(2, "semester") == "2 semesters"
    assert _word(1.0, "hour") == "hour"
    assert _word(4.0, "hour") == "hours"

    book = _book()
    text = "\n".join(
        str(cell.value)
        for sheet in book.worksheets
        for row in sheet.iter_rows()
        for cell in row
        if isinstance(cell.value, str)
    )
    assert "(s)" not in text.replace("Teacher(s)", "")     # only the column header keeps it
    assert "classs" not in text
    assert "1 classes" not in text and "3 class " not in text
    assert "contact 3 hours" not in text                   # no duplicated number


# ============= v2.2.0: free slots, load balancing, revisions, colour ========= #


def _entry(day, start, end, room, code, name, section, teacher, *, semester=1,
           credit=3, kind="theory", color="#a9d2e1", shift="morning", room_id=None,
           lab=0, students=30):
    rid = room_id if room_id is not None else hash(room) % 100 + 1
    return {
        "id": abs(hash((day, start, code, section, kind))) % 100000, "day": day,
        "start_time": start, "end_time": end, "shift": shift,
        "room_id": rid, "room_number": room, "room_label": f"B-{room}",
        "building_name": "B", "room_type": "Classroom", "capacity": 60,
        "course_id": abs(hash(code)) % 1000, "code": code, "course_name": name,
        "color": color, "department": "CS", "credit_hours": credit,
        "lab_credit_hours": lab, "section": section, "kind": kind,
        "semester": semester, "instructor": teacher, "num_students": students,
    }


_TWO_ROOMS = [
    {"id": 1, "room_number": "101", "label": "B-101", "capacity": 60, "room_type": "Classroom"},
    {"id": 2, "room_number": "102", "label": "B-102", "capacity": 40, "room_type": "Lab"},
]


def _cell_fill(sheet, row, column):
    cell = sheet.cell(row=row, column=column)
    if not cell.fill or cell.fill.fill_type != "solid":
        return ""
    return str(cell.fill.fgColor.rgb or "")


@pytest.mark.skipif(not openpyxl_available(), reason="openpyxl not installed")
def test_free_slots_shows_every_batch_an_open_slot():
    import io as _io

    from openpyxl import load_workbook

    from timetable.exporters import BUSY_FILL, FREE_FILL, build_workbook

    entries = [
        _entry(1, "08:30", "09:50", "101", "CS3009", "AI", "A", "Dr A", semester=1, room_id=1),
        _entry(1, "10:00", "11:20", "102", "CS4001", "ML", "B", "Dr B", semester=3, room_id=2),
    ]
    slots = [{"start": "08:30", "end": "09:50"}, {"start": "10:00", "end": "11:20"}]
    book = load_workbook(_io.BytesIO(build_workbook(
        entries, _TWO_ROOMS, days=2, slots=slots, show_summary=False, show_by_teacher=False,
        show_audit=False, show_dashboard=False, show_master_data=False, show_balance=False)))

    sheet = book["Free Slots"]
    grid = {}
    for row in range(1, sheet.max_row + 1):
        first = sheet.cell(row=row, column=1).value
        # only the per-batch matrix; the free-rooms table below also starts with
        # a weekday name and must not overwrite it
        if first in ("Monday", "Tuesday") and first not in grid:
            grid[first] = [
                (sheet.cell(row=row, column=2).value, _cell_fill(sheet, row, 2)),
                (sheet.cell(row=row, column=3).value, _cell_fill(sheet, row, 3)),
            ]
    # Section A is booked at 8:30 and free at 10:00
    assert grid["Monday"][0] == ("CS3009", BUSY_FILL)
    assert grid["Monday"][1] == ("free", FREE_FILL)
    # ...and the whole of Tuesday is open
    assert grid["Tuesday"][0] == ("free", FREE_FILL)
    assert grid["Tuesday"][1] == ("free", FREE_FILL)

    # the free-rooms table names a room the coordinator can actually book
    text = "\n".join(str(c.value) for r in sheet.iter_rows() for c in r if c.value is not None)
    assert "Free rooms at each slot" in text
    assert "B-102" in text


@pytest.mark.skipif(not openpyxl_available(), reason="openpyxl not installed")
def test_free_slots_turns_red_when_no_room_is_left():
    import io as _io

    from openpyxl import load_workbook

    from timetable.exporters import NO_ROOM_FILL, build_workbook

    # Section A has nothing at 10:00, but section B already holds the only room.
    entries = [
        _entry(1, "10:00", "11:20", "101", "CS4001", "ML", "B", "Dr B", semester=3, room_id=1),
        _entry(2, "08:30", "09:50", "101", "CS3009", "AI", "A", "Dr A", semester=1, room_id=1),
    ]
    slots = [{"start": "10:00", "end": "11:20"}]
    book = load_workbook(_io.BytesIO(build_workbook(
        entries, _TWO_ROOMS[:1], days=1, slots=slots, show_summary=False, show_by_teacher=False,
        show_audit=False, show_dashboard=False, show_master_data=False, show_balance=False)))

    sheet = book["Free Slots"]
    found = [
        (sheet.cell(row=r, column=2).value, _cell_fill(sheet, r, 2))
        for r in range(1, sheet.max_row + 1)
        if sheet.cell(row=r, column=1).value == "Monday"
    ]
    assert ("no room", NO_ROOM_FILL) in found


@pytest.mark.skipif(not openpyxl_available(), reason="openpyxl not installed")
def test_free_slots_and_balance_can_be_switched_off():
    book = _book(show_free_slots=False, show_balance=False)
    assert "Free Slots" not in book.sheetnames
    assert "Load Balancing" not in book.sheetnames


def test_load_balancing_suggests_a_move_only_to_someone_who_is_free():
    from timetable.reports import load_balance_suggestions

    entries = []
    # Dr A teaches six back-to-back 2-hour blocks: 12 h, well over the 20 h line
    # only when doubled, so push it past the threshold explicitly.
    for index in range(11):
        entries.append(_entry((index % 5) + 1, f"{8 + index // 5:02d}:00", f"{10 + index // 5:02d}:00",
                              "101", f"CS{index:04d}", f"Course {index}", "A", "Dr A"))
    entries.append(_entry(1, "14:00", "15:00", "102", "CS9000", "Light", "B", "Dr B"))

    report = load_balance_suggestions(entries)
    assert report["headers"][0] == "Over-loaded teacher"
    assert report["moves"] >= 1
    first = report["rows"][0]
    assert first[0] == "Dr A"
    assert first[6] == "Dr B"
    # the move must leave the receiver no busier than the giver
    giver_after, receiver_after = (float(part.strip().split()[0])
                                   for part in first[8].split("/"))
    assert receiver_after <= giver_after

    # Dr B is busy 2-3 PM, so nothing may be suggested at that time
    assert all(not (row[4] == "Monday" and row[5].startswith("2:00 PM")) for row in report["rows"])


def test_load_balancing_says_so_when_the_load_is_already_even():
    from timetable.reports import load_balance_suggestions

    entries = [
        _entry(1, "08:30", "09:50", "101", "CS3009", "AI", "A", "Dr A"),
        _entry(1, "10:00", "11:20", "102", "CS4001", "ML", "B", "Dr B"),
    ]
    report = load_balance_suggestions(entries)
    assert report["rows"] == []
    assert "0 moves suggested" in report["note"]


def test_the_balance_report_is_reachable_through_the_api(client):
    client.post("/api/timetable", json={"assignments": [
        {"day": 1, "start_time": "08:30", "end_time": "09:50", "room_id": 1,
         "course_id": 101, "section": "A", "shift": "morning"},
    ]})
    response = client.post("/api/report/balance", json={"days": 5})
    assert response.status_code == 200
    payload = response.get_json()
    assert payload["headers"][0] == "Over-loaded teacher"
    assert "moves" in payload
    assert client.post("/api/report/nonsense", json={}).status_code == 400


@pytest.mark.skipif(not openpyxl_available(), reason="openpyxl not installed")
def test_the_course_colour_follows_a_class_onto_every_sheet():
    """The colour a course wears on the grid is the colour it wears in Excel, so
    a subject can be followed down a page at a glance."""
    from timetable.exporters import _tint

    book = _book()
    expected = _tint("#a9d2e1", 0.74)

    semester = book["Semester 4"]
    header = _header_row(semester)
    assert _cell_fill(semester, header + 1, 3) == expected      # Course Title column

    summary = book["Summary"]
    summary_header = _header_row(summary, "Day")
    assert _cell_fill(summary, summary_header + 1, 9) == expected   # Course column

    master = book["Master Data"]
    painted = [
        _cell_fill(master, r, 3) for r in range(1, master.max_row + 1)
        if master.cell(row=r, column=2).value == "CHEM4134"
    ]
    assert expected in painted


@pytest.mark.skipif(not openpyxl_available(), reason="openpyxl not installed")
def test_the_summary_marks_evening_and_non_credited_rows():
    import io as _io

    from openpyxl import load_workbook

    from timetable.exporters import EVENING_FILL, WATCH_FILL, build_workbook

    entries = [
        _entry(1, "08:30", "09:50", "101", "CS3009", "AI", "A", "Dr A", shift="morning"),
        _entry(2, "18:00", "19:20", "101", "CS4001", "ML", "B", "Dr B", shift="evening"),
        _entry(3, "10:00", "11:00", "101", "CS5000", "Quran", "C", "Dr C", credit=0),
    ]
    book = load_workbook(_io.BytesIO(build_workbook(entries, _TWO_ROOMS[:1], days=5)))
    sheet = book["Summary"]
    header = _header_row(sheet, "Day")
    fills = {
        str(sheet.cell(row=r, column=2).value): _cell_fill(sheet, r, 2)
        for r in range(header + 1, header + 1 + len(entries))
    }
    assert fills["Evening"] == EVENING_FILL
    assert fills["Morning"] != EVENING_FILL

    # the 0-credit course is flagged in the C.Hrs column
    hours_fills = {
        sheet.cell(row=r, column=5).value: _cell_fill(sheet, r, 5)
        for r in range(header + 1, header + 1 + len(entries))
    }
    assert hours_fills[0] == WATCH_FILL


@pytest.mark.skipif(not openpyxl_available(), reason="openpyxl not installed")
def test_the_contents_page_is_colour_coded_by_kind_of_sheet():
    from timetable.exporters import TAB_COLOURS, _tint

    book = _book()
    contents = book["Contents"]
    header = _header_row(contents, "Sheet")
    rows = {
        contents.cell(row=r, column=1).value: _cell_fill(contents, r, 1)
        for r in range(header + 1, contents.max_row + 1)
        if contents.cell(row=r, column=1).value
    }
    assert rows["Semester 4"] == _tint(TAB_COLOURS["semester"][-6:], 0.82)
    assert rows["Monday"] == _tint(TAB_COLOURS["day"][-6:], 0.82)
    assert rows["Conflict Report"] == _tint(TAB_COLOURS["report"][-6:], 0.82)
    assert rows["Semester 4"] != rows["Monday"]      # the kinds really differ


def test_diff_summaries_catches_every_kind_of_change():
    from timetable.exporters import diff_summaries

    def row(code, name, section, day, start, room, teacher, kind="Theory"):
        return {"Code": code, "Course": name, "Section": section, "Type": kind,
                "Day": day, "Start": start, "End": "09:50 AM", "Room": room,
                "Teacher": teacher}

    before = [
        row("CS1", "One", "A", "Monday", "08:30 AM", "A-101", "Dr A"),
        row("CS2", "Two", "A", "Tuesday", "10:00 AM", "A-102", "Dr B"),
        row("CS3", "Three", "B", "Wednesday", "09:00 AM", "A-103", "Dr C"),
        row("CS4", "Four", "C", "Monday", "08:30 AM", "A-104", "Dr D"),
        row("CS4", "Four", "C", "Thursday", "11:00 AM", "A-104", "Dr D"),
        row("CS6", "Six", "E", "Friday", "09:00 AM", "A-106", "Dr F"),   # disappears
    ]
    after = [
        row("CS1", "One", "A", "Wednesday", "08:30 AM", "A-101", "Dr A"),   # moved
        row("CS2", "Two", "A", "Tuesday", "10:00 AM", "B-201", "Dr B"),     # re-roomed
        row("CS3", "Three", "B", "Wednesday", "09:00 AM", "A-103", "Dr Z"),  # re-taught
        row("CS4", "Four", "C", "Monday", "08:30 AM", "A-104", "Dr D"),     # untouched
        row("CS4", "Four", "C", "Friday", "11:00 AM", "A-104", "Dr D"),     # second meeting moved
        row("CS5", "Five", "D", "Thursday", "12:00 PM", "A-105", "Dr E"),   # added
    ]
    kinds = {kind for kind, _ in diff_summaries(before, after)}
    assert kinds == {"Added", "Removed", "Moved", "Room changed", "Teacher changed"}

    changes = dict((text.split(" — ")[0], kind) for kind, text in diff_summaries(before, after))
    # the two weekly meetings of CS4 are tracked separately: one kept, one moved
    cs4 = [text for kind, text in diff_summaries(before, after) if "CS4" in text]
    assert len(cs4) == 2
    assert "CS5 Five · Section D" in changes


def test_a_workbook_written_by_the_app_can_be_read_back_for_the_diff():
    import io as _io

    from timetable.exporters import build_workbook, diff_summaries, read_summary_rows

    data = build_workbook(_schedule_entries(), _SCHEDULE_ROOMS, days=5)
    rows = read_summary_rows(data)
    assert len(rows) == 3
    assert rows[0]["Day"] == "Monday"
    assert rows[0]["Course"] == "Special Paper - I"
    # comparing a workbook with itself reports no changes at all
    assert diff_summaries(rows, rows) == []
    # garbage in, empty list out - a bad file must never break an export
    assert read_summary_rows(b"not a workbook") == []


def test_revision_numbering_comes_from_what_is_already_in_the_folder(tmp_path):
    from timetable.filesystem import next_revision, revision_files

    assert next_revision(tmp_path, "Spring 2026")["revision"] == 1
    (tmp_path / "Spring 2026-rev1.xlsx").write_bytes(b"x" * 41832)
    (tmp_path / "Spring 2026-rev2.xlsx").write_bytes(b"y")
    (tmp_path / "Autumn 2025-rev7.xlsx").write_bytes(b"z")
    (tmp_path / "notes.txt").write_text("ignored")

    info = next_revision(tmp_path, "Spring 2026")
    assert info["revision"] == 3
    assert info["filename"] == "Spring 2026-rev3.xlsx"
    assert info["previous"] == "Spring 2026-rev2.xlsx"
    assert [item["revision"] for item in info["history"]] == [1, 2]
    assert info["history"][0]["size"] == "41 KB"

    # a different document in the same folder gets its own numbering
    assert next_revision(tmp_path, "Autumn 2025")["filename"] == "Autumn 2025-rev8.xlsx"
    assert [item["name"] for item in revision_files(tmp_path, "notes", ".txt")] == []
