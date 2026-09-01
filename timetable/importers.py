"""Bulk import of teachers, buildings, rooms, courses and sections from Excel.

The workbook this module reads is the same shape as the one
:func:`build_template` writes, so the usual workflow is:

    Import data  ->  Download template  ->  fill it in Excel  ->  Import

Every row is pushed through :class:`timetable.catalog.CatalogService`, which
means an import can never write anything the GUI would have refused: the same
validation, the same duplicate checks, the same referential guards.  Rows that
fail are reported back with their sheet name and row number instead of
aborting the whole file.
"""

from __future__ import annotations

import io
from typing import Any

from sqlalchemy import func, select

from .catalog import CatalogService
from .db import buildings, course_sections, courses, instructors, rooms
from .services import ValidationError

# sheet -> (column headings, example row)
TEMPLATE: dict[str, tuple[list[str], list[list[Any]]]] = {
    "Teachers": (
        ["Name", "Email", "Department", "Shift (morning/evening/both)"],
        [["Dr. Ayesha Khan", "a.khan@university.edu", "Computer Science", "both"]],
    ),
    "Buildings": (["Name"], [["A"], ["B"]]),
    "Rooms": (
        ["Room number", "Building", "Capacity", "Type (Classroom/Lab/Hall)"],
        [["301", "A", 60, "Classroom"], ["Lab-2", "B", 30, "Lab"]],
    ),
    "Courses": (
        ["Code", "Title", "Department", "Credit hours", "Colour (#RRGGBB)"],
        [["CS3009", "Artificial Intelligence", "Computer Science", 3, "#A9D2E1"]],
    ),
    "Sections": (
        ["Course code", "Section", "Teacher name or email"],
        [["CS3009", "A", "Dr. Ayesha Khan"], ["CS3009", "B", ""]],
    ),
}

SHEET_ORDER = ["Teachers", "Buildings", "Rooms", "Courses", "Sections"]


def openpyxl_available() -> bool:
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        return False
    return True


def build_template() -> bytes:
    """A ready-to-fill workbook: one sheet per entity, with example rows."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    workbook.remove(workbook.active)

    notes = workbook.create_sheet("Read me", 0)
    for row, line in enumerate(
        [
            "Automated Timetable Generator - import template",
            "",
            "1. Fill in the sheets you need. You may delete the example rows.",
            "2. Sheets you leave empty are simply skipped.",
            "3. Import order is fixed: Teachers, Buildings, Rooms, Courses, Sections.",
            "4. Existing records are matched and UPDATED, never duplicated:",
            "      Teachers  -> by name (or email)",
            "      Buildings -> by name",
            "      Rooms     -> by building + room number",
            "      Courses   -> by course code",
            "      Sections  -> by course code + section",
            "5. A row that fails validation is reported and skipped; the rest still import.",
        ],
        start=1,
    ):
        cell = notes.cell(row=row, column=1, value=line)
        if row == 1:
            cell.font = Font(bold=True, size=13, color="FF2B3465")
    notes.column_dimensions["A"].width = 92

    for name in SHEET_ORDER:
        headings, examples = TEMPLATE[name]
        sheet = workbook.create_sheet(name)
        for column, heading in enumerate(headings, start=1):
            cell = sheet.cell(row=1, column=column, value=heading)
            cell.font = Font(bold=True, color="FFFFFFFF")
            cell.fill = PatternFill("solid", fgColor="FF4C5CAF")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            sheet.column_dimensions[get_column_letter(column)].width = max(16, len(heading) + 4)
        for row_index, example in enumerate(examples, start=2):
            for column, value in enumerate(example, start=1):
                sheet.cell(row=row_index, column=column, value=value)
        sheet.freeze_panes = "A2"

    stream = io.BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _cell(row: tuple, index: int) -> str:
    if index >= len(row):
        return ""
    value = row[index]
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return str(value).strip()


class ImportReport:
    """Counters plus a per-row error list, ready to be shown in the UI."""

    def __init__(self) -> None:
        self.created: dict[str, int] = {}
        self.updated: dict[str, int] = {}
        self.skipped = 0
        self.errors: list[dict[str, Any]] = []

    def add(self, sheet: str, *, created: bool) -> None:
        target = self.created if created else self.updated
        target[sheet] = target.get(sheet, 0) + 1

    def fail(self, sheet: str, row: int, message: str) -> None:
        self.skipped += 1
        if len(self.errors) < 200:
            self.errors.append({"sheet": sheet, "row": row, "message": message})

    def as_dict(self) -> dict[str, Any]:
        total_created = sum(self.created.values())
        total_updated = sum(self.updated.values())
        parts = []
        for sheet in SHEET_ORDER:
            made, changed = self.created.get(sheet, 0), self.updated.get(sheet, 0)
            if made or changed:
                parts.append(f"{sheet}: {made} added, {changed} updated")
        return {
            "ok": not self.errors,
            "created": self.created,
            "updated": self.updated,
            "total_created": total_created,
            "total_updated": total_updated,
            "skipped": self.skipped,
            "errors": self.errors,
            "summary": " · ".join(parts) or "Nothing to import - every sheet was empty.",
        }


def import_workbook(catalog: CatalogService, data: bytes) -> dict[str, Any]:
    """Read an .xlsx file and upsert everything it contains."""
    from openpyxl import load_workbook

    try:
        workbook = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    except Exception as exc:  # openpyxl raises a zoo of exception types
        raise ValidationError(
            "That file could not be read as an Excel workbook (.xlsx). "
            f"Details: {exc}"
        ) from exc

    known = {name.strip().lower(): name for name in workbook.sheetnames}
    if not any(sheet.lower() in known for sheet in SHEET_ORDER):
        raise ValidationError(
            "No importable sheet found. The workbook needs at least one sheet named "
            + ", ".join(SHEET_ORDER)
            + ". Download the template to get the right layout."
        )

    report = ImportReport()
    engine = catalog.engine

    def rows_of(sheet_name: str):
        actual = known.get(sheet_name.lower())
        if not actual:
            return []
        sheet = workbook[actual]
        out = []
        for number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if number == 1:
                continue                     # heading
            if row is None or all(value in (None, "") for value in row):
                continue
            out.append((number, row))
        return out

    # -------------------------------------------------------------- teachers
    for number, row in rows_of("Teachers"):
        name = _cell(row, 0)
        email = _cell(row, 1)
        try:
            with engine.connect() as conn:
                found = conn.execute(
                    select(instructors.c.id).where(func.lower(instructors.c.name) == name.lower())
                ).first()
                if not found and email:
                    found = conn.execute(
                        select(instructors.c.id).where(func.lower(instructors.c.email) == email.lower())
                    ).first()
            catalog.save_instructor(
                {
                    "name": name,
                    "email": email,
                    "department": _cell(row, 2),
                    "shift": _cell(row, 3) or "both",
                },
                found.id if found else None,
            )
            report.add("Teachers", created=not found)
        except ValidationError as exc:
            report.fail("Teachers", number, str(exc))

    # ------------------------------------------------------------- buildings
    for number, row in rows_of("Buildings"):
        name = _cell(row, 0)
        try:
            with engine.connect() as conn:
                found = conn.execute(
                    select(buildings.c.id).where(func.lower(buildings.c.name) == name.lower())
                ).first()
            if found:
                report.add("Buildings", created=False)
                continue
            catalog.save_building({"name": name})
            report.add("Buildings", created=True)
        except ValidationError as exc:
            report.fail("Buildings", number, str(exc))

    # ----------------------------------------------------------------- rooms
    for number, row in rows_of("Rooms"):
        room_number = _cell(row, 0)
        building_name = _cell(row, 1)
        try:
            if not building_name:
                raise ValidationError("Building is required.")
            with engine.connect() as conn:
                building = conn.execute(
                    select(buildings.c.id).where(func.lower(buildings.c.name) == building_name.lower())
                ).first()
                found = None
                if building:
                    found = conn.execute(
                        select(rooms.c.id).where(
                            rooms.c.building_id == building.id,
                            func.lower(rooms.c.room_number) == room_number.lower(),
                        )
                    ).first()
            catalog.save_room(
                {
                    "room_number": room_number,
                    "building_name": building_name,
                    "capacity": _cell(row, 2) or 60,
                    "room_type": _cell(row, 3) or "Classroom",
                },
                found.id if found else None,
            )
            report.add("Rooms", created=not found)
        except ValidationError as exc:
            report.fail("Rooms", number, str(exc))

    # --------------------------------------------------------------- courses
    for number, row in rows_of("Courses"):
        code = _cell(row, 0)
        try:
            with engine.connect() as conn:
                found = conn.execute(
                    select(courses.c.id).where(func.upper(courses.c.code) == code.upper())
                ).first()
            catalog.save_course(
                {
                    "code": code,
                    "name": _cell(row, 1),
                    "department": _cell(row, 2),
                    "credit_hours": _cell(row, 3) or 3,
                    "color": _cell(row, 4),
                },
                found.id if found else None,
            )
            report.add("Courses", created=not found)
        except ValidationError as exc:
            report.fail("Courses", number, str(exc))

    # -------------------------------------------------------------- sections
    for number, row in rows_of("Sections"):
        code = _cell(row, 0)
        section = _cell(row, 1)
        teacher = _cell(row, 2)
        try:
            with engine.connect() as conn:
                course = conn.execute(
                    select(courses.c.id).where(func.upper(courses.c.code) == code.upper())
                ).first()
                if not course:
                    raise ValidationError(f"No course with code “{code}”. Add it on the Courses sheet first.")
                instructor_id = None
                if teacher:
                    match = conn.execute(
                        select(instructors.c.id).where(
                            (func.lower(instructors.c.name) == teacher.lower())
                            | (func.lower(instructors.c.email) == teacher.lower())
                        )
                    ).first()
                    if not match:
                        raise ValidationError(f"Unknown teacher “{teacher}”. Add them on the Teachers sheet first.")
                    instructor_id = match.id
                existing = conn.execute(
                    select(course_sections.c.section).where(
                        course_sections.c.course_id == course.id,
                        func.upper(course_sections.c.section) == section.upper(),
                    )
                ).first()
            catalog.save_section(course.id, {"section": section, "instructor_id": instructor_id})
            report.add("Sections", created=not existing)
        except ValidationError as exc:
            report.fail("Sections", number, str(exc))

    workbook.close()
    return report.as_dict()
