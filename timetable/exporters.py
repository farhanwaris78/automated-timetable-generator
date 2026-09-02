"""Excel (.xlsx) export.

Three layouts, all built from the same building blocks so a workbook always
looks like one document:

``book`` (the default)
    The **semester book**.  Every semester gets its own sheet drawn in the
    printed *Class Schedule* arrangement (metadata title block, one header
    row, rows grouped by day with the day cell merged and painted in a pastel
    band), and the roll-up sheets the previous releases had - Summary, one
    sheet per weekday, By Teacher, the three report sheets and the unscheduled
    list - are drawn in that same arrangement around them.  On top of those it
    adds a hyperlinked Contents page, a credit-hour audit, a charted dashboard
    and a master-data sheet.

``schedule``
    A single ``Class Schedule`` sheet - the one-page hand-out that matches the
    reference spreadsheet exactly.

``grid``
    The facilities view: room x time grids (one sheet per day, one per
    semester) with the same roll-up sheets around them.

Uses openpyxl, which is pure Python and therefore bundles cleanly into the
frozen executable.  If openpyxl is somehow unavailable the caller falls back
to CSV so the feature degrades instead of crashing.
"""

from __future__ import annotations

import io
from datetime import datetime
from typing import Any, Callable, Iterable, Sequence

from .services import WEEKDAYS, format_12h, to_minutes

HEADER_FILL = "FF4C5CAF"
BAND_FILL = "FFF2F4F9"
TITLE_FILL = "FFD9E2F3"
TOTALS_FILL = "FFE4EAF6"
SUBHEAD_FILL = "FFEFF3FA"

HEADING_COLOR = "FF1F3864"
TEXT_COLOR = "FF16223D"
NOTE_COLOR = "FF5B6479"
ALERT_COLOR = "FFB4380F"

# The default typewriter font for every exported workbook.  The user asked for
# "all text in Times New Roman", so this is the default; it can be changed in
# the Export & share dialog (and per request for the API).
DEFAULT_FONT_NAME = "Times New Roman"
DEFAULT_FONT_SIZE = 10


def openpyxl_available() -> bool:
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        return False
    return True


def _ink_for(hex_color: str) -> str:
    """Black or white text, whichever is readable on the given background."""
    value = (hex_color or "").lstrip("#")
    if len(value) != 6:
        return "FF000000"
    r, g, b = (int(value[i : i + 2], 16) for i in (0, 2, 4))
    return "FF16223D" if (0.299 * r + 0.587 * g + 0.114 * b) > 150 else "FFFFFFFF"


def _argb(hex_color: str, fallback: str = "FFDDDDDD") -> str:
    value = (hex_color or "").lstrip("#").upper()
    return f"FF{value}" if len(value) == 6 else fallback


def _class_label(entry: dict[str, Any]) -> str:
    code = entry.get("code") or ""
    lab = " [LAB]" if (entry.get("kind") == "lab") else ""
    return f"{code + ' · ' if code else ''}{entry['course_name']} ({entry['section']}){lab}"


def format_time_range(start: str, end: str) -> str:
    """12-hour range complete with AM/PM, e.g. \"2:30 PM - 4:00 PM\"."""
    def twelve(value: str) -> str:
        total = to_minutes(value)
        hour, minute = divmod(total, 60)
        suffix = "AM" if hour < 12 else "PM"
        display = hour % 12 or 12
        return f"{display}:{minute:02d} {suffix}"

    return f"{twelve(start)} - {twelve(end)}"


def _minutes(entry: dict[str, Any]) -> int:
    try:
        return to_minutes(str(entry.get("start_time") or "00:00"))
    except Exception:
        return 0


def _duration_hours(entry: dict[str, Any]) -> float:
    try:
        span = to_minutes(str(entry.get("end_time") or "")) - to_minutes(str(entry.get("start_time") or ""))
    except Exception:
        return 0.0
    return max(0, span) / 60.0


def _room_of(entry: dict[str, Any]) -> str:
    return str(entry.get("room_number") or entry.get("room_label") or "")


def _stamp() -> str:
    return datetime.now().strftime("%d %b %Y %H:%M")


def _word(count: Any, singular: str, plural: str | None = None) -> str:
    """Just the noun, correctly inflected: ``class`` / ``classes``."""
    try:
        number = int(count)
    except (TypeError, ValueError):
        return plural or f"{singular}s"
    if number == 1:
        return singular
    if plural:
        return plural
    if singular.endswith(("s", "x", "z", "ch", "sh")):
        return f"{singular}es"
    return f"{singular}s"


def _plural(count: Any, singular: str, plural: str | None = None) -> str:
    """``1 class`` / ``3 classes`` - the workbook is a document people read, so
    it should not say "1 class(es)"."""
    return f"{count} {_word(count, singular, plural)}"


# One pastel band per weekday, matching the reference Class Schedule sheet.
SCHEDULE_DAY_FILLS = {
    1: "FFA9C4E4",  # Monday   - blue
    2: "FFF2B8BC",  # Tuesday  - rose
    3: "FFB9ABD8",  # Wednesday- lavender
    4: "FFA9D6D6",  # Thursday - teal
    5: "FFC0B4DA",  # Friday   - lilac
    6: "FFD6C4E0",  # Saturday - mauve
    7: "FFE0D0D0",  # Sunday   - blush
}

# Alternating tints for the groupings that are not days (section / teacher).
GROUP_FILLS = ["FFEFF3FA", "FFF7F0F5", "FFEFF7F2", "FFFDF6EA", "FFF3F0FA"]

# A course with zero credit hours is reported as a non-credited course.  This
# is a shared label between the Excel and PDF schedules (and the only place
# the wording lives), so changing it here updates every export.
NON_CREDITED_LABEL = "non-credited course"

# Tab colours, so the sheet bar reads as a document rather than a pile of tabs.
# ARGB (with the alpha byte) - Excel treats a missing alpha as transparent and
# then draws no tab colour at all.
TAB_COLOURS = {
    "contents": "FF1F3864",
    "summary": "FF4C5CAF",
    "semester": "FF2E7D5B",
    "day": "FF7A6BB5",
    "teacher": "FFB5762E",
    "report": "FFB4380F",
    "extra": "FF5B6479",
}

# Status words -> tint.  Anything unrecognised is left unfilled.
STATUS_FILLS: tuple[tuple[str, str], ...] = (
    ("not scheduled", "FFF8D7DA"),
    ("over-loaded", "FFF8D7DA"),
    ("short", "FFFCE8D5"),
    ("under-used", "FFFCE8D5"),
    ("under-loaded", "FFFCE8D5"),
    ("extra", "FFFCE8D5"),
    ("error", "FFF8D7DA"),
    ("warning", "FFFDF1D6"),
    ("well used", "FFE6F2E6"),
    ("balanced", "FFE6F2E6"),
    ("complete", "FFE6F2E6"),
    ("ok", "FFE6F2E6"),
)


def _status_fill(value: Any) -> str:
    text = str(value or "").strip().lower()
    for needle, fill in STATUS_FILLS:
        if needle in text:
            return fill
    return ""


# --------------------------------------------------------------------------- #
# The printed "Class Schedule" column set
# --------------------------------------------------------------------------- #
# Every schedule-style sheet leads with the column its rows are grouped by and
# then carries the reference columns, so a reader can move between the
# Semester, weekday and teacher sheets without re-learning the layout.
SCHEDULE_COLUMNS = [
    "Course Code",
    "Course Title",
    "C.Hrs",
    "Total No.of Students",
    "Teacher's Name",
    "Time",
    "Room No",
]

# (kept as the public name the previous release used)
SCHEDULE_HEADERS = ["Days", *SCHEDULE_COLUMNS]
DAY_SHEET_HEADERS = ["Section", *SCHEDULE_COLUMNS]
TEACHER_SHEET_HEADERS = [
    "Teacher",
    "Day",
    "Course Code",
    "Course Title",
    "C.Hrs",
    "Total No.of Students",
    "Time",
    "Room No",
]

SUMMARY_HEADERS = [
    "Day", "Shift", "Semester", "Type", "C.Hrs", "Start", "End",
    "Code", "Course", "Section", "Teacher", "Room", "Students",
]

SUMMARY_WIDTHS = [12, 10, 10, 9, 8, 11, 11, 12, 30, 9, 26, 14, 10]
SCHEDULE_WIDTHS = [13, 14, 30, 18, 16, 24, 18, 10]
DAY_SHEET_WIDTHS = [18, 14, 30, 18, 16, 24, 18, 10]

# Columns whose text reads better left-aligned (long prose); everything else
# is centred, exactly like the reference spreadsheet.
LEFT_ALIGN_HEADERS = frozenset({
    "Course Title", "Teacher's Name", "Course", "Teacher", "Issue",
    "Room", "What is inside",
})

# A course-section whose contact hours are within this many hours of its
# planned credit hours counts as complete; slot lengths never land exactly on
# the hour, so a tighter tolerance would flag every row.
AUDIT_TOLERANCE_HOURS = 0.5

AUDIT_HEADERS = [
    "Semester", "Code", "Course", "Section", "Type", "Teacher",
    "Planned hrs / week", "Contact hrs on grid", "Difference", "Status",
]
AUDIT_WIDTHS = [11, 13, 32, 9, 9, 24, 17, 18, 12, 16]

UNPLACED_HEADERS = ["Semester", "Code", "Course", "Section", "Type", "Teacher", "Planned hrs / week"]
UNPLACED_WIDTHS = [11, 13, 32, 9, 9, 24, 17]

MASTER_COURSE_HEADERS = [
    "Semester", "Code", "Course", "Credit hrs", "Lab hrs", "Sections",
    "Teacher(s)", "Enrolled",
]
MASTER_TEACHER_HEADERS = ["Teacher", "Sections", "Classes / week", "Contact hrs / week", "Courses"]
MASTER_ROOM_HEADERS = [
    "Room", "Building", "Type", "Capacity", "Classes / week", "Contact hrs / week",
]


class _Theme:
    """Every style decision for one workbook lives here.

    Passing a single theme around (instead of re-creating Font objects in each
    sheet builder) is what keeps a 15-sheet workbook looking like one document:
    the same face, the same size, the same hairline border, the same palette.
    """

    def __init__(self, font_name: str = DEFAULT_FONT_NAME, font_size: int = DEFAULT_FONT_SIZE) -> None:
        from openpyxl.styles import Alignment, Border, Font, Side

        self.font_name = font_name
        self.font_size = font_size
        self._font_cls = Font
        self.thin = Side(style="thin", color="FFBFC5D8")
        self.border = Border(left=self.thin, right=self.thin, top=self.thin, bottom=self.thin)
        self.centre = Alignment(horizontal="center", vertical="center", wrap_text=True)
        self.left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        self.right = Alignment(horizontal="right", vertical="center", wrap_text=True)
        self.top_left = Alignment(horizontal="left", vertical="top", wrap_text=True)

    def font(self, *, size: int | None = None, bold: bool = False, color: str | None = None,
             italic: bool = False, underline: str | None = None):
        """One Font factory so a single ``font_name`` is used everywhere."""
        return self._font_cls(
            name=self.font_name,
            size=size if size is not None else self.font_size,
            bold=bold,
            italic=italic,
            underline=underline,
            color=color or "FF000000",
        )


def _fill(color: str):
    from openpyxl.styles import PatternFill

    return PatternFill("solid", fgColor=color)


def _hf(text: str) -> str:
    """Escape a string for an Excel header/footer (``&`` is a control char)."""
    return str(text or "").replace("&", "&&")


def _apply_theme(workbook, theme: _Theme) -> None:
    """Make the workbook's *default* font our chosen font too, so even cells
    we forget to style explicitly (or that a user adds later) match."""
    try:
        workbook._named_styles["Normal"].font = theme.font(size=theme.font_size)
    except Exception:
        pass


def _title_block(
    sheet,
    ncols: int,
    theme: _Theme,
    *,
    heading: str,
    institution: str = "",
    program: str = "",
    semester: str = "",
    commencement: str = "",
    extra: Sequence[str] = (),
) -> int:
    """Draw the reference metadata block and return the row for the header.

    The block is exactly the one on the printed Class Schedule: a filled title
    row, the institution, a ``Name of program`` / ``Semester`` pair, the
    commencement line, then any extra context lines (e.g. "12 class(es)").
    A blank spacer row always follows, so the header row is predictable.
    """
    from openpyxl.styles import Alignment

    ncols = max(2, int(ncols))

    def merge_row(row: int, value: str, *, fill: str = "", bold: bool = True,
                  size: int = 12, color: str = TEXT_COLOR, align: Alignment | None = None):
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
        cell = sheet.cell(row=row, column=1, value=value)
        cell.font = theme.font(bold=bold, size=size, color=color)
        cell.alignment = align or theme.centre
        if fill:
            for column in range(1, ncols + 1):
                sheet.cell(row=row, column=column).fill = _fill(fill)
        return cell

    row = 1
    merge_row(row, heading, fill=TITLE_FILL, bold=True, size=14, color=HEADING_COLOR)
    sheet.row_dimensions[row].height = 22
    row += 1
    if institution:
        merge_row(row, institution, bold=True, size=12, color=HEADING_COLOR)
        sheet.row_dimensions[row].height = 18
        row += 1

    left_label = f"Name of program: {program}" if program else ""
    right_label = f"Semester: {semester}" if semester else ""
    if left_label or right_label:
        half = max(2, ncols // 2)
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=half)
        lc = sheet.cell(row=row, column=1, value=left_label)
        lc.font = theme.font(size=11, color=TEXT_COLOR)
        lc.alignment = theme.left
        if right_label:
            sheet.merge_cells(start_row=row, start_column=half + 1, end_row=row, end_column=ncols)
            rc = sheet.cell(row=row, column=half + 1, value=right_label)
            rc.font = theme.font(size=11, color=TEXT_COLOR)
            rc.alignment = theme.left
        sheet.row_dimensions[row].height = 18
        row += 1

    if commencement:
        merge_row(row, f"Commencement of Classes: {commencement}", bold=False, size=11)
        sheet.row_dimensions[row].height = 18
        row += 1

    for line in extra:
        merge_row(row, line, bold=False, size=10, color=NOTE_COLOR, align=theme.left)
        sheet.row_dimensions[row].height = 15
        row += 1

    return row + 1          # one blank spacer row, then the header


def _draw_header(sheet, row: int, headers: Sequence[str], theme: _Theme, *, height: int = 30) -> None:
    for column, heading in enumerate(headers, start=1):
        cell = sheet.cell(row=row, column=column, value=heading)
        cell.fill = _fill(HEADER_FILL)
        cell.font = theme.font(bold=True, color="FFFFFFFF", size=11)
        cell.alignment = theme.centre
        cell.border = theme.border
    sheet.row_dimensions[row].height = height


def _set_widths(sheet, widths: Sequence[float]) -> None:
    from openpyxl.utils import get_column_letter

    for column, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width


def _finish_sheet(
    sheet,
    theme: _Theme,
    *,
    header_row: int,
    ncols: int,
    widths: Sequence[float],
    orientation: str = "landscape",
    tab: str = "",
    freeze: str | None = None,
    autofilter: str | None = None,
    gridlines: bool = False,
    footer_title: str = "",
    repeat_header: bool = True,
) -> None:
    """Column widths, print setup, tab colour, freeze panes - every sheet."""
    from openpyxl.utils import get_column_letter

    _set_widths(sheet, widths)
    sheet.sheet_view.showGridLines = gridlines
    if tab:
        sheet.sheet_properties.tabColor = TAB_COLOURS.get(tab, "FF5B6479")
    if freeze:
        sheet.freeze_panes = freeze
    if autofilter:
        sheet.auto_filter.ref = autofilter
    sheet.page_setup.orientation = orientation
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    if repeat_header and header_row:
        sheet.print_title_rows = f"{header_row}:{header_row}"
    sheet.page_margins.left = 0.4
    sheet.page_margins.right = 0.4
    sheet.page_margins.top = 0.6
    sheet.page_margins.bottom = 0.6
    sheet.oddFooter.left.text = _hf(footer_title)
    sheet.oddFooter.left.size = 8
    sheet.oddFooter.right.text = "Page &P of &N"
    sheet.oddFooter.right.size = 8
    _ = get_column_letter(max(1, ncols))


def _schedule_values(
    entry: dict[str, Any],
    *,
    flavour: str,
    non_credited_label: str = NON_CREDITED_LABEL,
    day: int | None = None,
    group: str = "",
) -> list[Any]:
    """One row of the printed Class Schedule, in the flavour's column order."""
    credits = int(entry.get("credit_hours") or 0)
    values = {
        "code": entry.get("code", ""),
        "name": entry.get("course_name", ""),
        "hours": non_credited_label if credits == 0 else str(credits),
        "students": entry.get("num_students", ""),
        "teacher": entry.get("instructor", "Unassigned"),
        "time": format_time_range(str(entry.get("start_time", "")), str(entry.get("end_time", ""))),
        "room": _room_of(entry),
        "kind": " [LAB]" if entry.get("kind") == "lab" else "",
    }
    day_name = WEEKDAYS[int(day if day is not None else entry.get("day") or 1) - 1]
    if flavour == "day":
        return [group, values["code"], f"{values['name']}{values['kind']}", values["hours"],
                values["students"], values["teacher"], values["time"], values["room"]]
    if flavour == "teacher":
        return [group, day_name, values["code"], f"{values['name']}{values['kind']}", values["hours"],
                values["students"], values["time"], values["room"]]
    return [day_name, values["code"], f"{values['name']}{values['kind']}", values["hours"],
            values["students"], values["teacher"], values["time"], values["room"]]


def _grouped_rows(
    sheet,
    start_row: int,
    entries: list[dict[str, Any]],
    *,
    theme: _Theme,
    headers: Sequence[str],
    flavour: str,
    group_of: Callable[[dict[str, Any]], tuple[Any, ...]],
    label_of: Callable[[tuple[Any, ...]], str],
    band_of: Callable[[int, tuple[Any, ...]], str],
    ncols: int,
    non_credited_label: str = NON_CREDITED_LABEL,
    empty_text: str = "No classes to show.",
    row_height: int = 26,
    label_size: int = 12,
) -> int:
    """Write the rows grouped by ``group_of``, merging the leading label cell.

    This is the arrangement the reference Class Schedule uses: the grouping key
    (day, section or teacher) is merged vertically over its block and painted a
    pastel band, the remaining columns hold the reference data.
    """
    ordered = sorted(entries, key=lambda e: (*group_of(e), _minutes(e), str(e.get("code") or "")))
    groups: list[tuple[tuple[Any, ...], list[dict[str, Any]]]] = []
    for entry in ordered:
        key = group_of(entry)
        if groups and groups[-1][0] == key:
            groups[-1][1].append(entry)
        else:
            groups.append((key, [entry]))

    if not groups:
        cell = sheet.cell(row=start_row, column=1, value=empty_text)
        cell.font = theme.font(italic=True, size=10, color=NOTE_COLOR)
        cell.alignment = theme.left
        return start_row + 1

    data_row = start_row
    for index, (key, block) in enumerate(groups):
        first = data_row
        band = band_of(index, key)
        day = int(key[0]) if flavour == "semester" and key and isinstance(key[0], int) else None
        for entry in block:
            credits = int(entry.get("credit_hours") or 0)
            values = _schedule_values(
                entry, flavour=flavour, non_credited_label=non_credited_label,
                day=day, group=label_of(key),
            )
            for column, value in enumerate(values, start=1):
                cell = sheet.cell(row=data_row, column=column, value=value)
                cell.border = theme.border
                cell.alignment = theme.left if headers[column - 1] in LEFT_ALIGN_HEADERS else theme.centre
                cell.fill = _fill(band)
                if column == 1:
                    continue
                if headers[column - 1] == "C.Hrs" and credits == 0:
                    cell.font = theme.font(bold=True, italic=True, color="FF9C4330",
                                           size=int(max(8, theme.font_size - 1)))
                else:
                    cell.font = theme.font(size=theme.font_size)
            sheet.row_dimensions[data_row].height = row_height
            data_row += 1

        if data_row - 1 > first:
            sheet.merge_cells(start_row=first, start_column=1, end_row=data_row - 1, end_column=1)
        label = sheet.cell(row=first, column=1, value=label_of(key))
        label.font = theme.font(bold=True, size=label_size, color=TEXT_COLOR)
        label.alignment = theme.centre
        for offset in range(first, data_row):
            sheet.cell(row=offset, column=1).fill = _fill(band)
            sheet.cell(row=offset, column=1).border = theme.border

        # A blank spacer row keeps the blocks visibly separated.
        data_row += 1

    return data_row - 1


def _totals_row(sheet, row: int, ncols: int, theme: _Theme, pairs: Sequence[tuple[str, Any]]) -> int:
    """A bold, filled strip of totals under a schedule table."""
    column = 1
    for label, value in pairs:
        cell = sheet.cell(row=row, column=column, value=f"{label}: {value}" if label else value)
        cell.font = theme.font(bold=True, size=10, color=TEXT_COLOR)
        cell.alignment = theme.centre
        cell.fill = _fill(TOTALS_FILL)
        cell.border = theme.border
        column += 1
    for rest in range(column, ncols + 1):
        cell = sheet.cell(row=row, column=rest)
        cell.fill = _fill(TOTALS_FILL)
        cell.border = theme.border
    sheet.row_dimensions[row].height = 20
    return row + 1


def _note(sheet, row: int, theme: _Theme, text: str, *, color: str = NOTE_COLOR, bold: bool = False) -> int:
    cell = sheet.cell(row=row, column=1, value=text)
    cell.font = theme.font(italic=not bold, bold=bold, size=9, color=color)
    cell.alignment = theme.left
    return row + 1


def _day_band(index: int, key: tuple[Any, ...]) -> str:
    day = int(key[0]) if key and isinstance(key[0], int) else 0
    return SCHEDULE_DAY_FILLS.get(day, GROUP_FILLS[index % len(GROUP_FILLS)])


def _cycle_band(index: int, key: tuple[Any, ...]) -> str:
    return GROUP_FILLS[index % len(GROUP_FILLS)]


def _semesters_in(entries: Iterable[dict[str, Any]]) -> list[int]:
    return sorted({int(e.get("semester") or 0) for e in entries if int(e.get("semester") or 0)})


def _section_label(entry: dict[str, Any]) -> str:
    sem = int(entry.get("semester") or 0)
    section = str(entry.get("section") or "").upper()
    prefix = f"Semester {sem}" if sem else "Semester —"
    return f"{prefix} · Section {section}" if section else prefix


def _week_totals(entries: list[dict[str, Any]]) -> dict[str, Any]:
    """Headline numbers for one set of rows (a week, a semester, a day…).

    ``credits`` counts each course-section once - a course scheduled as a
    lecture *and* a lab appears twice in ``entries`` but is only worth its
    credit hours once, plus the lab hours on top.
    """
    hours = round(sum(_duration_hours(e) for e in entries), 2)
    sections: dict[tuple[int, str], dict[str, int]] = {}
    for entry in entries:
        key = (int(entry.get("course_id") or 0), str(entry.get("section") or "").upper())
        bucket = sections.setdefault(key, {"credit": 0, "lab": 0})
        bucket["credit"] = int(entry.get("credit_hours") or 0)
        if entry.get("kind") == "lab":
            bucket["lab"] = int(entry.get("lab_credit_hours") or 0) or 1
    credits = sum(bucket["credit"] + bucket["lab"] for bucket in sections.values())
    return {
        "classes": len(entries),
        "hours": hours,
        "credits": credits,
        "teachers": len({str(e.get("instructor") or "Unassigned") for e in entries}),
        "sections": len(sections),
        "rooms": len({int(e.get("room_id") or 0) for e in entries}),
        "non_credited": sum(1 for e in entries if not int(e.get("credit_hours") or 0)),
    }


# --------------------------------------------------------------------------- #
# Sheet builders - one per kind of sheet, shared by every layout
# --------------------------------------------------------------------------- #
def _draw_semester_sheet(
    workbook,
    theme: _Theme,
    entries: list[dict[str, Any]],
    *,
    name: str,
    meta: dict[str, str],
    heading: str,
    orientation: str,
    non_credited_label: str,
    semester: str = "",
) -> dict[str, Any]:
    """One semester, in the printed Class Schedule arrangement."""
    sheet = workbook.create_sheet(name)
    totals = _week_totals(entries)
    extra = [
        f"{_plural(totals['classes'], 'class')} per week · {totals['hours']} contact "
        f"{_word(totals['hours'], 'hour')} · {_plural(totals['sections'], 'section')} · "
        f"{_plural(totals['teachers'], 'teacher')}"
    ]
    header_row = _title_block(
        sheet, len(SCHEDULE_HEADERS), theme,
        heading=heading, extra=extra,
        institution=meta.get("institution", ""),
        program=meta.get("program", ""),
        semester=semester or meta.get("semester", ""),
        commencement=meta.get("commencement", ""),
    )
    _draw_header(sheet, header_row, SCHEDULE_HEADERS, theme)
    last = _grouped_rows(
        sheet, header_row + 1, entries,
        theme=theme, headers=SCHEDULE_HEADERS, flavour="semester",
        group_of=lambda e: (int(e.get("day") or 0),),
        label_of=lambda key: WEEKDAYS[int(key[0]) - 1] if 1 <= int(key[0]) <= 7 else "—",
        band_of=_day_band, ncols=len(SCHEDULE_HEADERS),
        non_credited_label=non_credited_label,
        empty_text="Nothing is scheduled for this semester yet.",
    )
    _totals_row(
        sheet, last + 1, len(SCHEDULE_HEADERS), theme,
        [("Total", _plural(totals["classes"], "class")), ("Credit hours", totals["credits"]),
         ("Contact hrs / week", totals["hours"]), ("Sections", totals["sections"]),
         ("Teachers", totals["teachers"]), ("Non-credited", totals["non_credited"])],
    )
    _note(
        sheet, last + 3, theme,
        f"A course with 0 credit hours is written as \u201c{non_credited_label}\u201d. "
        f"Times are 12-hour with AM/PM.  Generated {_stamp()}.",
    )
    _finish_sheet(
        sheet, theme, header_row=header_row, ncols=len(SCHEDULE_HEADERS), widths=SCHEDULE_WIDTHS,
        orientation=orientation, tab="semester", freeze=f"B{header_row + 1}", footer_title=name,
    )
    return {"name": name, "classes": totals["classes"],
            "detail": f"{_plural(totals['sections'], 'section')} · "
                      f"{_plural(totals['teachers'], 'teacher')}"}


def _draw_day_sheet(
    workbook,
    theme: _Theme,
    entries: list[dict[str, Any]],
    *,
    day: int,
    meta: dict[str, str],
    heading: str,
    orientation: str,
    non_credited_label: str,
) -> dict[str, Any]:
    """One weekday, in the printed Class Schedule arrangement.

    Rows are grouped by *section* (semester + section) rather than by day,
    because the whole sheet is one day: a coordinator can read a batch's whole
    day straight down the page.
    """
    name = WEEKDAYS[day - 1]
    sheet = workbook.create_sheet(name)
    totals = _week_totals(entries)
    header_row = _title_block(
        sheet, len(DAY_SHEET_HEADERS), theme,
        heading=f"{heading} — {name}",
        extra=[f"{_plural(totals['classes'], 'class')} · {totals['hours']} contact "
               f"{_word(totals['hours'], 'hour')} · {_plural(totals['sections'], 'section')}"],
        institution=meta.get("institution", ""),
        program=meta.get("program", ""),
        semester=meta.get("semester", ""),
        commencement=meta.get("commencement", ""),
    )
    _draw_header(sheet, header_row, DAY_SHEET_HEADERS, theme)
    last = _grouped_rows(
        sheet, header_row + 1, entries,
        theme=theme, headers=DAY_SHEET_HEADERS, flavour="day",
        group_of=lambda e: (int(e.get("semester") or 0), str(e.get("section") or "").upper()),
        label_of=lambda key: (f"Semester {key[0]}" if key[0] else "Semester —") +
                             (f"\nSection {key[1]}" if key[1] else ""),
        band_of=_cycle_band, ncols=len(DAY_SHEET_HEADERS),
        non_credited_label=non_credited_label,
        empty_text=f"No classes are scheduled on {name}.",
        row_height=30,
    )
    if entries:
        _totals_row(
            sheet, last + 1, len(DAY_SHEET_HEADERS), theme,
            [("", f"{name} total"), ("Classes", totals["classes"]),
             ("Contact hours", totals["hours"]), ("Sections", totals["sections"]),
             ("Teachers", totals["teachers"]), ("Rooms used", totals["rooms"])],
        )
    _finish_sheet(
        sheet, theme, header_row=header_row, ncols=len(DAY_SHEET_HEADERS), widths=DAY_SHEET_WIDTHS,
        orientation=orientation, tab="day", freeze=f"B{header_row + 1}", footer_title=f"{heading} — {name}",
    )
    return {"name": name, "classes": totals["classes"],
            "detail": f"{_plural(totals['sections'], 'section')} · {_plural(totals['rooms'], 'room')}"}


def _draw_teacher_sheet(
    workbook,
    theme: _Theme,
    entries: list[dict[str, Any]],
    *,
    meta: dict[str, str],
    heading: str,
    orientation: str,
    non_credited_label: str,
) -> dict[str, Any]:
    """Every class per teacher, grouped (and merged) by teacher."""
    sheet = workbook.create_sheet("By Teacher")
    totals = _week_totals(entries)
    header_row = _title_block(
        sheet, len(TEACHER_SHEET_HEADERS), theme,
        heading=f"{heading} — By Teacher",
        extra=[f"{_plural(totals['teachers'], 'teacher')} · {_plural(totals['classes'], 'class')} · "
               f"{totals['hours']} contact {_word(totals['hours'], 'hour')}"],
        institution=meta.get("institution", ""),
        program=meta.get("program", ""),
        semester=meta.get("semester", ""),
        commencement=meta.get("commencement", ""),
    )
    _draw_header(sheet, header_row, TEACHER_SHEET_HEADERS, theme)
    last = _grouped_rows(
        sheet, header_row + 1, entries,
        theme=theme, headers=TEACHER_SHEET_HEADERS, flavour="teacher",
        group_of=lambda e: (str(e.get("instructor") or "Unassigned"),),
        label_of=lambda key: str(key[0]),
        band_of=_cycle_band, ncols=len(TEACHER_SHEET_HEADERS),
        non_credited_label=non_credited_label,
        empty_text="No classes are scheduled yet.",
    )
    _finish_sheet(
        sheet, theme, header_row=header_row, ncols=len(TEACHER_SHEET_HEADERS), widths=SCHEDULE_WIDTHS,
        orientation=orientation, tab="teacher", freeze=f"B{header_row + 1}", footer_title="By Teacher",
    )
    return {"name": "By Teacher", "classes": totals["classes"],
            "detail": _plural(totals["teachers"], "teacher")}


def _draw_summary_sheet(
    workbook,
    theme: _Theme,
    entries: list[dict[str, Any]],
    *,
    meta: dict[str, str],
    heading: str,
    orientation: str,
) -> dict[str, Any]:
    """The flat, filterable list of every class - the workbook's data sheet."""
    sheet = workbook.create_sheet("Summary")
    ordered = sorted(
        entries,
        key=lambda e: (int(e.get("day") or 0), _minutes(e), str(e.get("room_label") or "")),
    )
    totals = _week_totals(entries)
    header_row = _title_block(
        sheet, len(SUMMARY_HEADERS), theme,
        heading=f"{heading} — Summary",
        extra=[f"{_plural(totals['classes'], 'class')} · {totals['hours']} contact "
               f"{_word(totals['hours'], 'hour')} · {_plural(len(_semesters_in(entries)), 'semester')} · "
               "filter with the arrows in the header row"],
        institution=meta.get("institution", ""),
        program=meta.get("program", ""),
        semester=meta.get("semester", ""),
        commencement=meta.get("commencement", ""),
    )
    _draw_header(sheet, header_row, SUMMARY_HEADERS, theme, height=28)

    for offset, entry in enumerate(ordered):
        row = header_row + 1 + offset
        values = [
            WEEKDAYS[int(entry["day"]) - 1],
            (entry.get("shift") or "morning").capitalize(),
            int(entry.get("semester") or 0) or "-",
            "Lab" if entry.get("kind") == "lab" else "Theory",
            int(entry.get("credit_hours") or 0),
            format_12h(entry["start_time"]),
            format_12h(entry["end_time"]),
            entry.get("code", ""),
            entry["course_name"],
            entry.get("section", ""),
            entry.get("instructor", ""),
            entry.get("room_label", entry.get("room_id")),
            entry.get("num_students", ""),
        ]
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row=row, column=column, value=value)
            cell.border = theme.border
            cell.font = theme.font(size=theme.font_size)
            cell.alignment = theme.left if column in (9, 11, 12) else theme.centre
            if offset % 2 == 1:
                cell.fill = _fill(BAND_FILL)

    from openpyxl.utils import get_column_letter

    last_row = header_row + len(ordered)
    autofilter = f"A{header_row}:{get_column_letter(len(SUMMARY_HEADERS))}{max(header_row, last_row)}"
    _note(
        sheet, last_row + 2, theme,
        f"Generated {_stamp()} by Automated Timetable Generator.  "
        f"{_plural(totals['classes'], 'class')}, {totals['hours']} contact "
        f"{_word(totals['hours'], 'hour')} per week.",
    )
    _finish_sheet(
        sheet, theme, header_row=header_row, ncols=len(SUMMARY_HEADERS), widths=SUMMARY_WIDTHS,
        orientation=orientation, tab="summary", freeze=f"A{header_row + 1}",
        autofilter=autofilter, footer_title=f"{heading} — Summary",
    )
    return {"name": "Summary", "classes": totals["classes"],
            "detail": f"{_plural(len(_semesters_in(entries)), 'semester')}, "
                      "every class on one filterable sheet"}


def _draw_report_sheet(
    workbook,
    theme: _Theme,
    report: dict[str, Any],
    *,
    name: str,
    meta: dict[str, str],
    orientation: str,
) -> dict[str, Any]:
    """One of the three reports, with the same title block as everything else."""
    headers = list(report.get("headers") or [])
    if not headers:
        return {"name": name, "classes": 0, "detail": "empty"}
    sheet = workbook.create_sheet(name)
    ncols = len(headers)
    rows = report.get("rows") or []
    header_row = _title_block(
        sheet, ncols, theme,
        heading=str(report.get("title") or name),
        extra=[f"{_plural(len(rows), 'row')} · {report.get('note', '')}".strip(" ·"),
               f"Generated {_stamp()}"],
        institution=meta.get("institution", ""),
    )
    _draw_header(sheet, header_row, headers, theme, height=28)

    if not rows:
        cell = sheet.cell(row=header_row + 1, column=1, value="Nothing to report — no issues found.")
        cell.font = theme.font(italic=True, size=10, color=NOTE_COLOR)
        cell.alignment = theme.left

    status_column = next(
        (index for index, head in enumerate(headers, start=1)
         if str(head).strip().lower() in ("status", "load", "severity")),
        0,
    )
    for offset, row in enumerate(rows):
        target = header_row + 1 + offset
        for column, value in enumerate(row, start=1):
            cell = sheet.cell(row=target, column=column, value=value)
            cell.border = theme.border
            cell.font = theme.font(size=theme.font_size)
            cell.alignment = theme.left if column in (1, 6, 8) else theme.centre
            if offset % 2 == 1:
                cell.fill = _fill(BAND_FILL)
        if status_column:
            tint = _status_fill(row[status_column - 1])
            if tint:
                status = sheet.cell(row=target, column=status_column)
                status.fill = _fill(tint)
                status.font = theme.font(size=theme.font_size, bold=True)

    widths = [10, 10, 10, 13, 14, 10] if "Load" in headers else [12, 13, 12, 14, 12, 13, 12, 40]
    _finish_sheet(
        sheet, theme, header_row=header_row, ncols=ncols, widths=widths[:ncols] or [14] * ncols,
        orientation=orientation, tab="report", freeze=f"A{header_row + 1}",
        autofilter=f"A{header_row}:{chr(64 + ncols)}{max(header_row, header_row + len(rows))}"
        if ncols <= 26 else None,
        footer_title=str(report.get("title") or name),
    )
    return {"name": name, "classes": len(rows),
            "detail": str(report.get("note") or _plural(len(rows), "row"))}


def _draw_unscheduled_sheet(
    workbook,
    theme: _Theme,
    unscheduled: list[dict[str, Any]],
    *,
    meta: dict[str, str],
    heading: str,
    orientation: str,
) -> dict[str, Any]:
    sheet = workbook.create_sheet("Unscheduled")
    ncols = len(UNPLACED_HEADERS)
    header_row = _title_block(
        sheet, ncols, theme,
        heading=f"{heading} — Unscheduled",
        extra=[f"{_plural(len(unscheduled), 'class')} still "
               f"{'needs' if len(unscheduled) == 1 else 'need'} a slot on the grid"],
        institution=meta.get("institution", ""),
    )
    _draw_header(sheet, header_row, UNPLACED_HEADERS, theme, height=28)
    if not unscheduled:
        cell = sheet.cell(row=header_row + 1, column=1,
                          value="Everything in the catalogue is scheduled. 🎉")
        cell.font = theme.font(italic=True, size=10, color="FF2E7D5B")
        cell.alignment = theme.left
    for offset, item in enumerate(unscheduled):
        row = header_row + 1 + offset
        values = [
            int(item.get("semester") or 0) or "-",
            item.get("code", ""),
            item.get("course_name", ""),
            item.get("section", ""),
            "Lab" if item.get("kind") == "lab" else "Theory",
            item.get("instructor", ""),
            int(item.get("hours") or 0) or "-",
        ]
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row=row, column=column, value=value)
            cell.border = theme.border
            cell.font = theme.font(size=theme.font_size)
            cell.alignment = theme.left if column in (3, 6) else theme.centre
            if offset % 2 == 1:
                cell.fill = _fill(BAND_FILL)
    _finish_sheet(
        sheet, theme, header_row=header_row, ncols=ncols, widths=UNPLACED_WIDTHS,
        orientation=orientation, tab="report", freeze=f"A{header_row + 1}",
        footer_title=f"{heading} — Unscheduled",
    )
    return {"name": "Unscheduled", "classes": len(unscheduled),
            "detail": "classes the catalogue expects but the grid does not have"}


def _audit_rows(entries: list[dict[str, Any]], unscheduled: list[dict[str, Any]]) -> list[list[Any]]:
    """Planned credit hours per course-section vs. contact hours actually on the grid."""
    planned: dict[tuple[Any, ...], dict[str, Any]] = {}

    def key_of(course_id: Any, section: Any, kind: Any) -> tuple[Any, ...]:
        return (int(course_id or 0), str(section or "").upper(), str(kind or "theory"))

    for entry in entries:
        key = key_of(entry.get("course_id"), entry.get("section"), entry.get("kind"))
        bucket = planned.setdefault(key, {
            "semester": int(entry.get("semester") or 0),
            "code": entry.get("code", ""),
            "name": entry.get("course_name", ""),
            "section": entry.get("section", ""),
            "kind": entry.get("kind") or "theory",
            "teacher": entry.get("instructor", ""),
            "planned": 0.0,
            "actual": 0.0,
        })
        credits = int(entry.get("credit_hours") or 0)
        bucket["planned"] = float(
            int(entry.get("lab_credit_hours") or 0) if bucket["kind"] == "lab" else credits
        )
        bucket["actual"] += _duration_hours(entry)

    for item in unscheduled:
        key = key_of(item.get("course_id"), item.get("section"), item.get("kind"))
        planned.setdefault(key, {
            "semester": int(item.get("semester") or 0),
            "code": item.get("code", ""),
            "name": item.get("course_name", ""),
            "section": item.get("section", ""),
            "kind": item.get("kind") or "theory",
            "teacher": item.get("instructor", ""),
            "planned": float(int(item.get("hours") or 0)),
            "actual": 0.0,
        })

    rows: list[list[Any]] = []
    for key in sorted(planned, key=lambda k: (planned[k]["semester"], str(planned[k]["code"]),
                                              str(planned[k]["name"]), planned[k]["section"])):
        bucket = planned[key]
        planned_hours = round(bucket["planned"], 2)
        actual_hours = round(bucket["actual"], 2)
        difference = round(actual_hours - planned_hours, 2)
        if actual_hours == 0:
            status = "Not scheduled"
        elif planned_hours == 0:
            status = "Non-credited"
        elif abs(difference) <= AUDIT_TOLERANCE_HOURS:
            status = "Complete"
        elif difference < 0:
            status = f"Short {abs(difference):g} h"
        else:
            status = f"Extra {difference:g} h"
        rows.append([
            bucket["semester"] or "-",
            bucket["code"],
            bucket["name"],
            bucket["section"],
            "Lab" if bucket["kind"] == "lab" else "Theory",
            bucket["teacher"] or "Unassigned",
            planned_hours,
            actual_hours,
            difference,
            status,
        ])
    return rows


def _draw_audit_sheet(
    workbook,
    theme: _Theme,
    entries: list[dict[str, Any]],
    unscheduled: list[dict[str, Any]],
    *,
    meta: dict[str, str],
    heading: str,
    orientation: str,
) -> dict[str, Any]:
    """Planned credit hours vs. what the grid actually delivers, per section."""
    rows = _audit_rows(entries, unscheduled)
    sheet = workbook.create_sheet("Credit Hour Audit")
    ncols = len(AUDIT_HEADERS)
    bad = sum(1 for row in rows if row[9] not in ("Complete", "Non-credited"))
    header_row = _title_block(
        sheet, ncols, theme,
        heading=f"{heading} — Credit Hour Audit",
        extra=[f"{_plural(len(rows), 'course-section')} · "
               f"{bad} {'needs' if bad == 1 else 'need'} attention"],
        institution=meta.get("institution", ""),
    )
    _draw_header(sheet, header_row, AUDIT_HEADERS, theme, height=28)
    for offset, row in enumerate(rows):
        target = header_row + 1 + offset
        for column, value in enumerate(row, start=1):
            cell = sheet.cell(row=target, column=column, value=value)
            cell.border = theme.border
            cell.font = theme.font(size=theme.font_size)
            cell.alignment = theme.left if column in (3, 6) else theme.centre
            if offset % 2 == 1:
                cell.fill = _fill(BAND_FILL)
        tint = _status_fill(row[9])
        if tint:
            status = sheet.cell(row=target, column=10)
            status.fill = _fill(tint)
            status.font = theme.font(size=theme.font_size, bold=True)
    _note(
        sheet, header_row + len(rows) + 2, theme,
        "“Planned” is the catalogue credit hours (lab hours for a lab row); “contact hours on grid” is the "
        "time the grid actually books.  A difference of more than half an hour is flagged, so a 3-credit "
        "course that only reaches 80 minutes a week shows up as Short 1.67 h.  Non-credited courses "
        "(0 credit hours) are never flagged.",
    )
    _finish_sheet(
        sheet, theme, header_row=header_row, ncols=ncols, widths=AUDIT_WIDTHS,
        orientation=orientation, tab="extra", freeze=f"A{header_row + 1}",
        autofilter=f"A{header_row}:J{max(header_row, header_row + len(rows))}",
        footer_title="Credit Hour Audit",
    )
    return {"name": "Credit Hour Audit", "classes": len(rows),
            "detail": f"{bad} {'needs' if bad == 1 else 'need'} attention"}


def _draw_dashboard_sheet(
    workbook,
    theme: _Theme,
    entries: list[dict[str, Any]],
    rooms: list[dict[str, Any]],
    unscheduled: list[dict[str, Any]],
    *,
    meta: dict[str, str],
    heading: str,
    orientation: str,
    days: int,
    slots: list[dict[str, str]] | None,
    shift: str,
) -> dict[str, Any]:
    """A one-page overview: the headline numbers plus two bar charts."""
    from .reports import conflict_report, room_utilisation, teacher_workload

    utilisation = room_utilisation(entries, rooms, days=days, slots=slots, shift=shift)
    workload = teacher_workload(entries)
    clashes = conflict_report(entries)
    totals = _week_totals(entries)
    errors = sum(1 for row in clashes.get("rows") or [] if row[0] == "Error")
    warnings = sum(1 for row in clashes.get("rows") or [] if row[0] == "Warning")

    sheet = workbook.create_sheet("Dashboard")
    header_row = _title_block(
        sheet, 6, theme,
        heading=f"{heading} — Dashboard",
        extra=[f"Week at a glance · generated {_stamp()}"],
        institution=meta.get("institution", ""),
        program=meta.get("program", ""),
        semester=meta.get("semester", ""),
        commencement=meta.get("commencement", ""),
    )
    _draw_header(sheet, header_row, ["Measure", "This week", "", "", "", ""], theme, height=22)

    kpis: list[tuple[str, Any]] = [
        ("Scheduled classes / week", totals["classes"]),
        ("Contact hours / week", totals["hours"]),
        ("Credit hours on the grid", totals["credits"]),
        ("Course sections taught", totals["sections"]),
        ("Teachers on the grid", totals["teachers"]),
        ("Rooms used", totals["rooms"]),
        ("Semesters covered", len(_semesters_in(entries))),
        ("Non-credited classes", totals["non_credited"]),
        ("Classes still unscheduled", len(unscheduled)),
        ("Clashes to fix (errors)", errors),
        ("Warnings", warnings),
    ]
    row = header_row + 1
    for offset, (label, value) in enumerate(kpis):
        lc = sheet.cell(row=row, column=1, value=label)
        lc.font = theme.font(size=theme.font_size, bold=True)
        lc.alignment = theme.left
        lc.border = theme.border
        vc = sheet.cell(row=row, column=2, value=value)
        vc.font = theme.font(size=theme.font_size)
        vc.alignment = theme.centre
        vc.border = theme.border
        if str(label).startswith("Clashes") and errors:
            vc.fill = _fill("FFF8D7DA")
            vc.font = theme.font(size=theme.font_size, bold=True, color="FF9C0006")
        elif offset % 2 == 1:
            lc.fill = _fill(BAND_FILL)
            vc.fill = _fill(BAND_FILL)
        row += 1

    # Chart source data lives out of the way (columns H/I) and is hidden, so the
    # charts have something numeric to read without cluttering the page.
    from openpyxl.chart import BarChart, Reference

    data_row = header_row
    sheet.cell(row=data_row, column=8, value="Room").font = theme.font(size=9, bold=True)
    sheet.cell(row=data_row, column=9, value="Utilisation %").font = theme.font(size=9, bold=True)
    rooms_chart = [row for row in (utilisation.get("rows") or [])][:10]
    for offset, source in enumerate(rooms_chart, start=1):
        sheet.cell(row=data_row + offset, column=8, value=str(source[0]))
        try:
            sheet.cell(row=data_row + offset, column=9, value=float(str(source[6]).replace("%", "")))
        except ValueError:
            sheet.cell(row=data_row + offset, column=9, value=0)

    workload_row = data_row + len(rooms_chart) + 3
    sheet.cell(row=workload_row, column=8, value="Teacher").font = theme.font(size=9, bold=True)
    sheet.cell(row=workload_row, column=9, value="Contact hours").font = theme.font(size=9, bold=True)
    teachers_chart = [row for row in (workload.get("rows") or [])][:10]
    for offset, source in enumerate(teachers_chart, start=1):
        sheet.cell(row=workload_row + offset, column=8, value=str(source[0]))
        try:
            sheet.cell(row=workload_row + offset, column=9, value=float(source[2]))
        except (TypeError, ValueError):
            sheet.cell(row=workload_row + offset, column=9, value=0)

    if rooms_chart:
        chart = BarChart()
        chart.type = "bar"
        chart.style = 10
        chart.title = "Room utilisation (%)"
        chart.y_axis.title = "Utilisation %"
        chart.height = 8
        chart.width = 17
        chart.visible_cells_only = False      # the source columns are hidden
        chart.add_data(
            Reference(sheet, min_col=9, min_row=data_row, max_row=data_row + len(rooms_chart)),
            titles_from_data=True,
        )
        chart.set_categories(
            Reference(sheet, min_col=8, min_row=data_row + 1, max_row=data_row + len(rooms_chart))
        )
        sheet.add_chart(chart, "D4")

    if teachers_chart:
        chart = BarChart()
        chart.type = "bar"
        chart.style = 12
        chart.title = "Teacher contact hours per week"
        chart.y_axis.title = "Hours"
        chart.height = 8
        chart.width = 17
        chart.visible_cells_only = False      # the source columns are hidden
        chart.add_data(
            Reference(sheet, min_col=9, min_row=workload_row, max_row=workload_row + len(teachers_chart)),
            titles_from_data=True,
        )
        chart.set_categories(
            Reference(sheet, min_col=8, min_row=workload_row + 1,
                      max_row=workload_row + len(teachers_chart))
        )
        sheet.add_chart(chart, "D21")

    sheet.column_dimensions["H"].hidden = True
    sheet.column_dimensions["I"].hidden = True
    _finish_sheet(
        sheet, theme, header_row=header_row, ncols=6, widths=[34, 14, 3, 12, 12, 12],
        orientation=orientation, tab="extra", freeze=None, repeat_header=False,
        footer_title="Dashboard",
    )
    return {"name": "Dashboard", "classes": totals["classes"],
            "detail": f"{_plural(errors, 'error')}, {_plural(warnings, 'warning')}, "
                      f"{_plural(len(unscheduled), 'class')} unscheduled"}


def _draw_master_sheet(
    workbook,
    theme: _Theme,
    entries: list[dict[str, Any]],
    rooms: list[dict[str, Any]],
    *,
    meta: dict[str, str],
    heading: str,
    orientation: str,
) -> dict[str, Any]:
    """The catalogue behind the grid: courses, teachers and rooms in one place."""
    sheet = workbook.create_sheet("Master Data")
    ncols = max(len(MASTER_COURSE_HEADERS), len(MASTER_TEACHER_HEADERS), len(MASTER_ROOM_HEADERS))
    header_row = _title_block(
        sheet, ncols, theme,
        heading=f"{heading} — Master Data",
        extra=["Every course, teacher and room the grid is built from"],
        institution=meta.get("institution", ""),
    )

    def table(row: int, title: str, headers: Sequence[str], rows: Sequence[Sequence[Any]],
              widths: Sequence[float]) -> int:
        cell = sheet.cell(row=row, column=1, value=title)
        cell.font = theme.font(bold=True, size=12, color=HEADING_COLOR)
        cell.alignment = theme.left
        row += 1
        _draw_header(sheet, row, headers, theme, height=24)
        for offset, values in enumerate(rows):
            target = row + 1 + offset
            for column, value in enumerate(values, start=1):
                c = sheet.cell(row=target, column=column, value=value)
                c.border = theme.border
                c.font = theme.font(size=theme.font_size)
                c.alignment = theme.left if column in (2, 3, 5) else theme.centre
                if offset % 2 == 1:
                    c.fill = _fill(BAND_FILL)
        return row + len(rows) + 3

    # ---- courses ---------------------------------------------------------- #
    courses: dict[tuple[Any, ...], dict[str, Any]] = {}
    for entry in entries:
        key = (int(entry.get("course_id") or 0), str(entry.get("code") or ""), str(entry.get("course_name") or ""))
        bucket = courses.setdefault(key, {
            "semester": int(entry.get("semester") or 0),
            "credit": int(entry.get("credit_hours") or 0),
            "lab": int(entry.get("lab_credit_hours") or 0),
            "sections": set(),
            "teachers": set(),
            "students": 0,
        })
        bucket["lab"] = max(bucket["lab"], int(entry.get("lab_credit_hours") or 0))
        bucket["credit"] = max(bucket["credit"], int(entry.get("credit_hours") or 0))
        bucket["sections"].add(str(entry.get("section") or "").upper())
        bucket["teachers"].add(str(entry.get("instructor") or "Unassigned"))
        bucket["students"] = max(bucket["students"], int(entry.get("num_students") or 0))
    course_rows = [
        [bucket["semester"] or "-", key[1], key[2], bucket["credit"], bucket["lab"] or "-",
         ", ".join(sorted(s for s in bucket["sections"] if s)),
         ", ".join(sorted(bucket["teachers"])), bucket["students"]]
        for key, bucket in sorted(courses.items(), key=lambda item: (item[1]["semester"], item[0][1], item[0][2]))
    ]
    row = table(header_row, "Courses", MASTER_COURSE_HEADERS, course_rows, SUMMARY_WIDTHS)

    # ---- teachers --------------------------------------------------------- #
    teachers: dict[str, dict[str, Any]] = {}
    for entry in entries:
        name = str(entry.get("instructor") or "Unassigned")
        bucket = teachers.setdefault(name, {"sections": set(), "classes": 0, "hours": 0.0, "courses": set()})
        bucket["sections"].add(f"{entry.get('code') or entry.get('course_name')}-{str(entry.get('section')).upper()}")
        bucket["classes"] += 1
        bucket["hours"] += _duration_hours(entry)
        bucket["courses"].add(str(entry.get("code") or entry.get("course_name") or ""))
    teacher_rows = [
        [name, len(b["sections"]), b["classes"], round(b["hours"], 2), len(b["courses"])]
        for name, b in sorted(teachers.items(), key=lambda item: -item[1]["hours"])
    ]
    row = table(row, "Teachers", MASTER_TEACHER_HEADERS, teacher_rows, [26, 10, 14, 18, 10])

    # ---- rooms ------------------------------------------------------------ #
    used: dict[int, dict[str, Any]] = {}
    for entry in entries:
        rid = int(entry.get("room_id") or 0)
        bucket = used.setdefault(rid, {"classes": 0, "hours": 0.0,
                                       "building": entry.get("building_name") or "",
                                       "type": entry.get("room_type") or "",
                                       "capacity": int(entry.get("capacity") or 0),
                                       "label": entry.get("room_label") or entry.get("room_number") or rid})
        bucket["classes"] += 1
        bucket["hours"] += _duration_hours(entry)
    for room in rooms:
        used.setdefault(int(room["id"]), {
            "classes": 0, "hours": 0.0,
            "building": room.get("building") or room.get("building_name") or "",
            "type": room.get("room_type") or room.get("type") or "",
            "capacity": int(room.get("capacity") or 0),
            "label": room.get("label") or room.get("room_number") or room["id"],
        })
    room_rows = [
        [b["label"], b["building"], b["type"], b["capacity"], b["classes"], round(b["hours"], 2)]
        for b in sorted(used.values(), key=lambda item: str(item["label"]))
    ]
    table(row, "Rooms", MASTER_ROOM_HEADERS, room_rows, [16, 18, 14, 10, 14, 18])

    _finish_sheet(
        sheet, theme, header_row=0, ncols=ncols,
        widths=[14, 16, 32, 12, 12, 16, 24, 12],
        orientation=orientation, tab="extra", freeze=None, repeat_header=False,
        footer_title="Master Data",
    )
    return {"name": "Master Data", "classes": len(course_rows),
            "detail": f"{_plural(len(course_rows), 'course')}, "
                      f"{_plural(len(teacher_rows), 'teacher')}, {_plural(len(room_rows), 'room')}"}


def _draw_contents_sheet(
    workbook,
    theme: _Theme,
    index: list[dict[str, Any]],
    *,
    meta: dict[str, str],
    heading: str,
    orientation: str,
    layout: str,
    total_classes: int,
) -> None:
    """A hyperlinked table of contents: what is in the workbook, and how big."""
    from openpyxl.worksheet.hyperlink import Hyperlink

    sheet = workbook.create_sheet("Contents", 0)
    ncols = 3
    lines = [f"{_plural(len(index), 'sheet')} · {_plural(total_classes, 'class')} on the grid · "
             f"layout: {layout}"]
    header_row = _title_block(
        sheet, ncols, theme,
        heading=heading,
        extra=lines + [f"Generated {_stamp()} by Automated Timetable Generator"],
        institution=meta.get("institution", ""),
        program=meta.get("program", ""),
        semester=meta.get("semester", ""),
        commencement=meta.get("commencement", ""),
    )
    _draw_header(sheet, header_row, ["Sheet", "What is inside", "Classes"], theme, height=24)

    row = header_row + 1
    for offset, item in enumerate(index):
        name = str(item.get("name") or "")
        link = sheet.cell(row=row, column=1, value=name)
        try:
            link.hyperlink = Hyperlink(ref=link.coordinate, location=f"'{name}'!A1", display=name)
        except Exception:  # pragma: no cover - very old openpyxl
            pass
        link.font = theme.font(size=theme.font_size, color="FF1F4E9C", underline="single")
        link.alignment = theme.left
        link.border = theme.border
        for column, value in enumerate([item.get("detail", ""), item.get("classes", "")], start=2):
            cell = sheet.cell(row=row, column=column, value=value)
            cell.font = theme.font(size=theme.font_size)
            cell.alignment = theme.left if column == 2 else theme.centre
            cell.border = theme.border
        if offset % 2 == 1:
            for column in range(1, ncols + 1):
                sheet.cell(row=row, column=column).fill = _fill(BAND_FILL)
        row += 1

    _note(
        sheet, row + 1, theme,
        "Click a sheet name to jump to it.  Every sheet prints on one page wide, repeats its header row on "
        "each printed page and carries page numbers in the footer.",
    )
    _finish_sheet(
        sheet, theme, header_row=header_row, ncols=ncols, widths=[24, 66, 10],
        orientation=orientation, tab="contents", freeze=f"A{header_row + 1}",
        repeat_header=False, footer_title="Contents",
    )


# --------------------------------------------------------------------------- #
# The workbook
# --------------------------------------------------------------------------- #
def build_workbook(
    entries: list[dict[str, Any]],
    rooms: list[dict[str, Any]],
    *,
    days: int = 5,
    slots: list[dict[str, str]] | None = None,
    shift: str = "all",
    title: str = "University Timetable",
    unscheduled: list[dict[str, Any]] | None = None,
    font_name: str = DEFAULT_FONT_NAME,
    font_size: int = DEFAULT_FONT_SIZE,
    orientation: str = "landscape",
    institution: str = "",
    term: str = "",
    show_summary: bool = True,
    show_by_teacher: bool = True,
    show_unscheduled: bool = True,
    show_semesters: bool = True,
    layout: str = "book",
    program: str = "",
    commencement: str = "",
    semester: str = "",
    non_credited_label: str = NON_CREDITED_LABEL,
    show_audit: bool = True,
    show_dashboard: bool = True,
    show_master_data: bool = True,
    contents: bool = True,
) -> bytes:
    """Render the timetable into an .xlsx workbook and return the bytes.

    ``layout``
        ``book`` (default)   - the semester book: a Contents page, Summary, one
                               Class Schedule sheet **per semester**, one per
                               weekday, By Teacher, Credit Hour Audit, a
                               charted Dashboard, the three report sheets,
                               Master Data and Unscheduled.  Every sheet uses
                               the printed Class Schedule arrangement.
        ``schedule``         - a single "Class Schedule" sheet: a metadata title
                               block (institution, program, semester,
                               commencement) and rows grouped by day with the
                               reference columns Days / Course Code / Course
                               Title / C.Hrs / Students / Teacher / Time /
                               Room No.
        ``grid``             - one room x time sheet per day plus day x section
                               grids per semester (the facilities view), with
                               the same roll-up sheets around them.

    Courses with ``credit_hours == 0`` are written as **non-credited course**.
    ``font_name`` is applied to *every* cell in the workbook, so the export
    reads consistently in Excel, LibreOffice and Google Sheets (Times New
    Roman by default).
    """
    from openpyxl import Workbook

    if layout == "schedule":
        return build_class_schedule_workbook(
            entries,
            days=days,
            title=title,
            font_name=font_name,
            font_size=font_size,
            orientation=orientation,
            institution=institution,
            term=term,
            program=program,
            commencement=commencement,
            semester=semester,
            non_credited_label=non_credited_label,
        )

    theme = _Theme(font_name, font_size)
    unscheduled = unscheduled or []
    if shift and shift != "all":
        entries = [e for e in entries if (e.get("shift") or "morning") == shift]

    # Derive the slot list from the data when the caller did not supply one.
    if not slots:
        seen: dict[str, str] = {}
        for entry in entries:
            seen[entry["start_time"]] = entry["end_time"]
        slots = [{"start": s, "end": seen[s]} for s in sorted(seen, key=to_minutes)]

    used_rooms = {e["room_id"] for e in entries}
    grid_rooms = [r for r in rooms if r["id"] in used_rooms] or rooms[:12]
    day_count = max(1, min(7, int(days or 1)))

    heading = str(title or "University Timetable")
    if term:
        heading = f"{heading} {term}"
    meta = {
        "institution": institution,
        "program": program,
        "semester": semester,
        "commencement": commencement,
    }

    workbook = Workbook()
    workbook.remove(workbook.active)
    _apply_theme(workbook, theme)
    workbook.properties.title = heading
    workbook.properties.subject = "University timetable"
    workbook.properties.creator = "Automated Timetable Generator"
    workbook.properties.description = (
        f"{heading}. {len(entries)} scheduled class(es). Generated {_stamp()}."
    )
    workbook.properties.keywords = "timetable, schedule, semester, university"

    index: list[dict[str, Any]] = []

    # ------------------------------------------------------------ summary --- #
    if show_summary:
        index.append(_draw_summary_sheet(workbook, theme, entries, meta=meta,
                                         heading=heading, orientation=orientation))

    semesters = _semesters_in(entries) if show_semesters else []

    if layout == "grid":
        index.extend(
            _draw_grid_day_sheets(workbook, theme, entries, grid_rooms, slots,
                                  days=day_count, shift=shift, heading=heading,
                                  meta=meta, orientation=orientation, institution=institution)
        )
        if semesters:
            index.extend(
                _draw_grid_semester_sheets(workbook, theme, entries, slots, semesters,
                                           days=day_count, heading=heading, meta=meta,
                                           orientation=orientation, institution=institution)
            )
    else:
        # ---- one Class Schedule sheet per semester ------------------------ #
        if semesters:
            for number in semesters:
                block = [e for e in entries if int(e.get("semester") or 0) == number]
                index.append(_draw_semester_sheet(
                    workbook, theme, block,
                    name=f"Semester {number}",
                    meta=meta,
                    heading=f"{heading} — Semester {number}",
                    orientation=orientation,
                    non_credited_label=non_credited_label,
                    semester=str(number),
                ))
        else:
            index.append(_draw_semester_sheet(
                workbook, theme, entries,
                name="All Classes",
                meta=meta,
                heading=f"{heading} — All Classes",
                orientation=orientation,
                non_credited_label=non_credited_label,
            ))

        # ---- one Class Schedule sheet per weekday ------------------------- #
        for day in range(1, day_count + 1):
            index.append(_draw_day_sheet(
                workbook, theme, [e for e in entries if int(e.get("day") or 0) == day],
                day=day, meta=meta, heading=heading, orientation=orientation,
                non_credited_label=non_credited_label,
            ))

    # ----------------------------------------------------------- by teacher -- #
    if show_by_teacher:
        index.append(_draw_teacher_sheet(workbook, theme, entries, meta=meta,
                                         heading=heading, orientation=orientation,
                                         non_credited_label=non_credited_label))

    # --------------------------------------------------------------- extras -- #
    if layout != "grid" and show_audit:
        index.append(_draw_audit_sheet(workbook, theme, entries, unscheduled, meta=meta,
                                       heading=heading, orientation=orientation))
    if layout != "grid" and show_dashboard:
        index.append(_draw_dashboard_sheet(workbook, theme, entries, rooms, unscheduled,
                                           meta=meta, heading=heading, orientation=orientation,
                                           days=day_count, slots=slots, shift=shift))

    # -------------------------------------------------------------- reports -- #
    from .reports import conflict_report, room_utilisation, teacher_workload

    index.append(_draw_report_sheet(
        workbook, theme, room_utilisation(entries, rooms, days=day_count, slots=slots, shift=shift),
        name="Room Utilisation", meta=meta, orientation=orientation,
    ))
    index.append(_draw_report_sheet(
        workbook, theme, teacher_workload(entries), name="Teacher Workload",
        meta=meta, orientation=orientation,
    ))
    index.append(_draw_report_sheet(
        workbook, theme, conflict_report(entries), name="Conflict Report",
        meta=meta, orientation=orientation,
    ))

    if layout != "grid" and show_master_data:
        index.append(_draw_master_sheet(workbook, theme, entries, rooms, meta=meta,
                                        heading=heading, orientation=orientation))

    if unscheduled and show_unscheduled:
        index.append(_draw_unscheduled_sheet(workbook, theme, unscheduled, meta=meta,
                                             heading=heading, orientation=orientation))

    if contents:
        _draw_contents_sheet(workbook, theme, index, meta=meta, heading=heading,
                             orientation=orientation, layout=layout, total_classes=len(entries))

    # Finally, sweep every sheet and force our chosen font onto any cell that
    # is still carrying the workbook template default ("Calibri").  We style
    # the header / title / grid cells explicitly, but body cells created by
    # ``cell(value=...)`` inherit openpyxl's default font; without this sweep
    # those would leak the template font into the export and break the
    # "all text in Times New Roman" requirement.
    for worksheet in workbook.worksheets:
        for row_cells in worksheet.iter_rows():
            for cell in row_cells:
                if cell.value is not None and cell.font.name != font_name:
                    cell.font = theme.font()

    stream = io.BytesIO()
    workbook.save(stream)
    return stream.getvalue()


# --------------------------------------------------------------------------- #
# Grid layout: room x time sheets (the facilities view)
# --------------------------------------------------------------------------- #
def _draw_grid_day_sheets(
    workbook,
    theme: _Theme,
    entries: list[dict[str, Any]],
    grid_rooms: list[dict[str, Any]],
    slots: list[dict[str, str]],
    *,
    days: int,
    shift: str,
    heading: str,
    meta: dict[str, str],
    orientation: str,
    institution: str,
) -> list[dict[str, Any]]:
    from openpyxl.utils import get_column_letter

    made: list[dict[str, Any]] = []
    for day in range(1, days + 1):
        name = WEEKDAYS[day - 1]
        sheet = workbook.create_sheet(name)
        day_entries = [e for e in entries if int(e["day"]) == day]

        sheet["A1"] = f"{heading} — {name}"
        sheet["A1"].font = theme.font(bold=True, size=14, color="FF2B3465")
        span = max(2, len(slots) + 1)
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
        sheet["A1"].alignment = theme.left
        label = "All shifts" if shift == "all" else f"{shift.capitalize()} shift"
        subrow = f"{label} · {_plural(len(day_entries), 'class')}"
        if institution or meta.get("institution"):
            subrow = f"{institution or meta.get('institution')} · {subrow}".strip(" ·")
        sheet.cell(row=2, column=1, value=subrow).font = theme.font(italic=True, size=10, color=NOTE_COLOR)

        header_row = 4
        head = sheet.cell(row=header_row, column=1, value="Room / Time")
        head.fill = _fill(HEADER_FILL)
        head.font = theme.font(bold=True, color="FFFFFFFF", size=11)
        head.alignment = theme.centre
        head.border = theme.border
        for index, slot in enumerate(slots, start=2):
            cell = sheet.cell(row=header_row, column=index,
                              value=f"{format_12h(slot['start'])}\n{format_12h(slot['end'])}")
            cell.fill = _fill(HEADER_FILL)
            cell.font = theme.font(bold=True, color="FFFFFFFF", size=11)
            cell.alignment = theme.centre
            cell.border = theme.border

        lookup = {(e["start_time"], e["room_id"]): e for e in day_entries}
        for offset, room in enumerate(grid_rooms):
            row_index = header_row + 1 + offset
            room_cell = sheet.cell(
                row=row_index, column=1,
                value=f"{room.get('label') or room['room_number']}  ({room.get('capacity', '')} seats)",
            )
            room_cell.font = theme.font(bold=True, size=10)
            room_cell.alignment = theme.left
            room_cell.border = theme.border
            room_cell.fill = _fill(BAND_FILL)

            for column, slot in enumerate(slots, start=2):
                cell = sheet.cell(row=row_index, column=column)
                cell.border = theme.border
                cell.alignment = theme.centre
                entry = lookup.get((slot["start"], room["id"]))
                if not entry:
                    continue
                cell.value = f"{_class_label(entry)}\n{entry.get('instructor') or ''}"
                cell.fill = _fill(_argb(entry.get("color")))
                cell.font = theme.font(size=9, color=_ink_for(entry.get("color") or ""))

        sheet.column_dimensions["A"].width = 26
        for column in range(2, len(slots) + 2):
            sheet.column_dimensions[get_column_letter(column)].width = 24
        for offset in range(len(grid_rooms)):
            sheet.row_dimensions[header_row + 1 + offset].height = 34
        sheet.row_dimensions[header_row].height = 30
        _finish_sheet(sheet, theme, header_row=header_row, ncols=len(slots) + 1,
                      widths=[26] + [24] * len(slots), orientation=orientation, tab="day",
                      freeze=f"B{header_row + 1}", gridlines=True, footer_title=f"{heading} — {name}",
                      repeat_header=False)
        made.append({"name": name, "classes": len(day_entries),
                     "detail": f"room x time grid · {_plural(len(grid_rooms), 'room')}"})
    return made


def _draw_grid_semester_sheets(
    workbook,
    theme: _Theme,
    entries: list[dict[str, Any]],
    slots: list[dict[str, str]],
    semesters: list[int],
    *,
    days: int,
    heading: str,
    meta: dict[str, str],
    orientation: str,
    institution: str,
) -> list[dict[str, Any]]:
    from openpyxl.utils import get_column_letter

    made: list[dict[str, Any]] = []
    for number in semesters:
        sem_entries = [e for e in entries if int(e.get("semester") or 0) == number]
        sections = sorted({str(e["section"]).upper() for e in sem_entries})
        name = f"Semester {number}"
        sheet = workbook.create_sheet(name)

        sheet["A1"] = f"{heading} — Semester {number}"
        sheet["A1"].font = theme.font(bold=True, size=14, color="FF2B3465")
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(2, len(slots) + 1))
        sheet["A1"].alignment = theme.left
        teachers_here = len({e.get("instructor") for e in sem_entries})
        subrow = (f"{_plural(len(sem_entries), 'class')} · {_plural(len(sections), 'section')} · "
                  f"{_plural(teachers_here, 'teacher')}")
        if institution or meta.get("institution"):
            subrow = f"{institution or meta.get('institution')} · {subrow}".strip(" ·")
        sheet.cell(row=2, column=1, value=subrow).font = theme.font(italic=True, size=10, color=NOTE_COLOR)

        header_row = 4
        head = sheet.cell(row=header_row, column=1, value="Day / Section")
        head.fill = _fill(HEADER_FILL)
        head.font = theme.font(bold=True, color="FFFFFFFF", size=11)
        head.alignment = theme.centre
        head.border = theme.border
        for index, slot in enumerate(slots, start=2):
            cell = sheet.cell(row=header_row, column=index,
                              value=f"{format_12h(slot['start'])}\n{format_12h(slot['end'])}")
            cell.fill = _fill(HEADER_FILL)
            cell.font = theme.font(bold=True, color="FFFFFFFF", size=11)
            cell.alignment = theme.centre
            cell.border = theme.border

        lookup = {(int(e["day"]), str(e["section"]).upper(), e["start_time"]): e for e in sem_entries}
        row_index = header_row
        for day in range(1, days + 1):
            for section in sections:
                row_index += 1
                label = sheet.cell(row=row_index, column=1, value=f"{WEEKDAYS[day - 1]} · Section {section}")
                label.font = theme.font(bold=True, size=10)
                label.alignment = theme.left
                label.border = theme.border
                label.fill = _fill(BAND_FILL)
                for column, slot in enumerate(slots, start=2):
                    cell = sheet.cell(row=row_index, column=column)
                    cell.border = theme.border
                    cell.alignment = theme.centre
                    entry = lookup.get((day, section, slot["start"]))
                    if not entry:
                        continue
                    cell.value = (
                        f"{_class_label(entry)}\n{entry.get('instructor') or ''}\n"
                        f"{entry.get('room_label') or ''}"
                    )
                    cell.fill = _fill(_argb(entry.get("color")))
                    cell.font = theme.font(size=9, color=_ink_for(entry.get("color") or ""))
                sheet.row_dimensions[row_index].height = 40

        _finish_sheet(sheet, theme, header_row=header_row, ncols=len(slots) + 1,
                      widths=[26] + [26] * len(slots), orientation=orientation, tab="semester",
                      freeze=f"B{header_row + 1}", gridlines=True, footer_title=name,
                      repeat_header=False)
        made.append({"name": name, "classes": len(sem_entries),
                     "detail": f"day x section grid · {_plural(len(sections), 'section')}"})
    return made


# --------------------------------------------------------------------------- #
# Class Schedule (the printed-classroom-timetable layout)
# --------------------------------------------------------------------------- #
def build_class_schedule_workbook(
    entries: list[dict[str, Any]],
    *,
    days: int = 5,
    title: str = "Class Schedule",
    font_name: str = DEFAULT_FONT_NAME,
    font_size: int = DEFAULT_FONT_SIZE,
    orientation: str = "landscape",
    institution: str = "",
    term: str = "",
    program: str = "",
    commencement: str = "",
    semester: str = "",
    non_credited_label: str = NON_CREDITED_LABEL,
) -> bytes:
    """Render the timetable as one printed-style "Class Schedule" sheet.

    Mirrors the reference spreadsheet: a merged metadata block up top
    (institution, program, semester, commencement) followed by a header row and
    rows grouped by day, where the Day cell is merged vertically over that
    day's classes and painted a pastel band.  A course whose credit hours are
    zero is written as ``non_credited_label`` (default **non-credited course**).
    """
    from openpyxl import Workbook

    theme = _Theme(font_name, font_size)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Class Schedule"
    _apply_theme(workbook, theme)
    workbook.properties.title = f"{title} {term}".strip()
    workbook.properties.creator = "Automated Timetable Generator"

    ncols = len(SCHEDULE_HEADERS)
    heading = str(title or "Class Schedule")
    if term:
        heading = f"{heading} {term}"
    header_row = _title_block(
        sheet, ncols, theme,
        heading=heading, institution=institution, program=program,
        semester=semester, commencement=commencement,
    )
    _draw_header(sheet, header_row, SCHEDULE_HEADERS, theme)

    ordered = sorted(entries, key=lambda e: (int(e["day"]), _minutes(e), str(e.get("code", ""))))
    days_present = sorted({int(e["day"]) for e in entries if 1 <= int(e["day"]) <= 7})
    if not days_present:
        days_present = list(range(1, min(7, max(1, days)) + 1))

    data_row = header_row + 1
    for day in days_present[:7]:
        day_entries = [e for e in ordered if int(e["day"]) == day]
        if not day_entries:
            continue
        start_row = data_row
        band = SCHEDULE_DAY_FILLS.get(day, "FFDDDDE8")

        for entry in day_entries:
            credits = int(entry.get("credit_hours") or 0)
            values = [
                WEEKDAYS[day - 1],
                entry.get("code", ""),
                entry.get("course_name", ""),
                non_credited_label if credits == 0 else str(credits),
                entry.get("num_students", ""),
                entry.get("instructor", "Unassigned"),
                format_time_range(str(entry.get("start_time", "")), str(entry.get("end_time", ""))),
                _room_of(entry),
            ]
            for column, value in enumerate(values, start=1):
                cell = sheet.cell(row=data_row, column=column, value=value)
                cell.border = theme.border
                cell.alignment = theme.left if column in (3, 6) else theme.centre
                if column == 4 and credits == 0:
                    cell.font = theme.font(bold=True, italic=True, color="FF9C4330",
                                           size=int(max(8, font_size - 1)))
                else:
                    cell.font = theme.font(size=font_size)
            for column in range(1, ncols + 1):
                sheet.cell(row=data_row, column=column).fill = _fill(band)
            sheet.row_dimensions[data_row].height = 26
            data_row += 1

        # Merge the day cell over its block and re-centre.
        if data_row - 1 > start_row:
            sheet.merge_cells(start_row=start_row, start_column=1, end_row=data_row - 1, end_column=1)
        day_cell = sheet.cell(row=start_row, column=1)
        day_cell.font = theme.font(bold=True, size=12, color=TEXT_COLOR)
        day_cell.alignment = theme.centre

        # A blank spacer row keeps the day blocks visibly separated.
        data_row += 1

    _finish_sheet(
        sheet, theme, header_row=header_row, ncols=ncols,
        widths=[12, 14, 30, 18, 16, 24, 18, 10],
        orientation=orientation, freeze=f"B{header_row + 1}", footer_title="Class Schedule",
    )

    # Norm the font on any cell we might have skipped.
    for row_cells in sheet.iter_rows():
        for cell in row_cells:
            if cell.value is not None and cell.font.name != font_name:
                cell.font = theme.font()

    stream = io.BytesIO()
    workbook.save(stream)
    return stream.getvalue()


# --------------------------------------------------------------------------- #
# CSV
# --------------------------------------------------------------------------- #
CSV_HEADERS = [
    "Day",
    "Shift",
    "Start",
    "End",
    "Room",
    "Building",
    "Room type",
    "Capacity",
    "Code",
    "Course",
    "Section",
    "Kind",
    "Teacher",
    "Semester",
    "Credit hours",
    "Students",
]


def _csv_rows(entries: list[dict[str, Any]]) -> list[list[Any]]:
    def sort_key(entry: dict[str, Any]) -> tuple[int, int]:
        return int(entry.get("day") or 0), _minutes(entry)

    rows: list[list[Any]] = []
    for entry in sorted(entries, key=sort_key):
        day = int(entry.get("day") or 0)
        rows.append([
            WEEKDAYS[day - 1] if 1 <= day <= len(WEEKDAYS) else day,
            str(entry.get("shift") or "morning").title(),
            format_12h(str(entry.get("start_time") or "")),
            format_12h(str(entry.get("end_time") or "")),
            entry.get("room_label") or entry.get("room_id") or "",
            entry.get("building_name") or entry.get("building") or "",
            entry.get("room_type") or "",
            entry.get("capacity") or "",
            entry.get("code") or "",
            entry.get("course_name") or "",
            entry.get("section") or "",
            "Lab" if entry.get("kind") == "lab" else "Theory",
            entry.get("instructor") or "Unassigned",
            entry.get("semester") or "",
            int(entry.get("credit_hours") or 0),
            entry.get("num_students") or 0,
        ])
    return rows


def _csv_bytes(headers: Sequence[str], rows: Iterable[Sequence[Any]]) -> bytes:
    import csv

    buffer = io.StringIO(newline="")
    writer = csv.writer(buffer, lineterminator="\r\n", quoting=csv.QUOTE_ALL)
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)
    return "\ufeff".encode("utf-8") + buffer.getvalue().encode("utf-8")


def build_csv(entries: list[dict[str, Any]]) -> bytes:
    """Render the timetable as UTF-8 CSV bytes.

    Written server-side (rather than stitched together in JavaScript) so the
    CSV can be saved straight into the project folder, contains exactly the
    same columns as the workbook, and is correctly quoted for Excel.  A UTF-8
    BOM is included so Excel opens accented names correctly on Windows.
    """
    return _csv_bytes(CSV_HEADERS, _csv_rows(entries))


def build_csv_bundle(
    entries: list[dict[str, Any]],
    *,
    unscheduled: list[dict[str, Any]] | None = None,
    shift: str = "all",
) -> bytes:
    """A .zip holding one CSV per workbook sheet.

    Same content as the Excel book, but in a form a script, a database loader
    or a colleague without Excel can read: ``timetable.csv`` (everything),
    ``semester-<n>.csv`` per semester, ``<weekday>.csv`` per weekday,
    ``by-teacher.csv``, ``credit-hour-audit.csv`` and ``unscheduled.csv``.
    """
    import zipfile

    if shift and shift != "all":
        entries = [e for e in entries if (e.get("shift") or "morning") == shift]
    unscheduled = unscheduled or []

    files: dict[str, bytes] = {"timetable.csv": build_csv(entries)}

    for number in _semesters_in(entries):
        block = [e for e in entries if int(e.get("semester") or 0) == number]
        files[f"semester-{number}.csv"] = build_csv(block)

    for day in range(1, 8):
        block = [e for e in entries if int(e.get("day") or 0) == day]
        if block:
            files[f"{WEEKDAYS[day - 1].lower()}.csv"] = build_csv(block)

    files["by-teacher.csv"] = _csv_bytes(
        ["Teacher", "Day", "Start", "End", "Code", "Course", "Section", "Kind", "Room", "Semester"],
        [
            [
                e.get("instructor") or "Unassigned",
                WEEKDAYS[int(e.get("day") or 1) - 1],
                format_12h(str(e.get("start_time") or "")),
                format_12h(str(e.get("end_time") or "")),
                e.get("code") or "",
                e.get("course_name") or "",
                e.get("section") or "",
                "Lab" if e.get("kind") == "lab" else "Theory",
                e.get("room_label") or "",
                e.get("semester") or "",
            ]
            for e in sorted(entries, key=lambda item: (str(item.get("instructor") or ""),
                                                       int(item.get("day") or 0), _minutes(item)))
        ],
    )

    files["credit-hour-audit.csv"] = _csv_bytes(AUDIT_HEADERS, _audit_rows(entries, unscheduled))

    if unscheduled:
        files["unscheduled.csv"] = _csv_bytes(
            UNPLACED_HEADERS,
            [
                [
                    int(item.get("semester") or 0) or "-",
                    item.get("code", ""),
                    item.get("course_name", ""),
                    item.get("section", ""),
                    "Lab" if item.get("kind") == "lab" else "Theory",
                    item.get("instructor", ""),
                    int(item.get("hours") or 0) or "-",
                ]
                for item in unscheduled
            ],
        )

    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as archive:
        for name, data in files.items():
            archive.writestr(name, data)
    return buffer.getvalue()
